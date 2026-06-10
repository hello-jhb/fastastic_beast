from __future__ import annotations

from pathlib import Path
import json
import re
import pandas as pd
import openpyxl

from metric_catalog import load_metric_catalog


UPLOAD_DIR = Path("uploads")
REPOSITORY_DIR = Path("repository")


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def is_numeric(value):
    return isinstance(value, (int, float)) and not pd.isna(value)


def is_datetime(value):
    """Excel date cells come through as datetime.datetime via openpyxl."""
    import datetime as _dt
    return isinstance(value, (_dt.datetime, _dt.date))


def is_value_for_metric(value, metric_unit: str | None = None):
    """
    Accept numeric values universally, plus datetime values when the metric
    explicitly expects a date (unit == 'date'). Used so Phase 1.5a's date
    anchor metrics (Purchase Date, Exit Date) can find date cells.
    """
    if is_numeric(value):
        return True
    if metric_unit == "date" and is_datetime(value):
        return True
    return False


def normalize_text(value):
    return clean_text(value).lower()


def cell_address(row, col):
    return openpyxl.utils.get_column_letter(col) + str(row)


# =============================================================================
# Sheet priority taxonomy
# =============================================================================
# Defines which sheets to look at first when extracting metrics, AND which
# sheets to skip entirely. Lower tier = higher priority.
#
# Rationale: institutional models put deal-level numbers on summary sheets.
# Cash flow projections live on dedicated CF/proforma sheets. CapEx and debt
# have their own. Waterfall comes last. Backup/comps/historicals/sensitivities
# are noise for first-pass extraction and often cause WRONG matches (e.g. an
# "NOI" label in a market comp sheet overwrites the real underwriting NOI).
#
# When the same metric is found on multiple sheets, the match from the
# lower-tier (higher priority) sheet wins, regardless of confidence.

# Tier definitions use generic categorical keywords that should apply to ANY
# institutional real estate model regardless of property type or sponsor template.
#
# Discipline applied when picking keywords:
#   - Generic real estate / finance terminology only (not property-type-specific)
#   - No sheet name from any specific file we've tested
#   - Each keyword must plausibly appear in models for: multifamily, office,
#     industrial, retail, hotel, mixed-use, dev, value-add, core, portfolio
#   - Substring match — keep keywords short enough that variations of the
#     same concept are caught (e.g. "summary" catches "Deal Summary",
#     "Investment Summary", "Summary & Assumptions", etc.)

SHEET_PRIORITY_TIERS: list[tuple[int, list[str]]] = [
    # Tier 1 — deal-level summary (highest priority)
    # An analyst reading any UW model goes here first.
    (1, [
        "summary",          # catches: Deal Summary, Investment Summary, Summary & Assumptions, Executive Summary
        "one pager", "one-pager", "onepager",
        "overview", "snapshot", "dashboard", "highlights",
        "general info",     # catches: General Information
        "key metric", "key uw", "key input", "key assumption",
    ]),
    # Tier 2 — cash flow projections (the proforma)
    # Time-series of revenue, expenses, NOI by year/quarter/month.
    (2, [
        "cash flow", "cashflow",
        "proforma", "pro forma", "pro-forma",
        "p&l", "p & l", "pnl",
        "operating statement", "operating income",
        "annual cf", "monthly cf", "quarterly cf",
        "annual cfs", "monthly cfs",
    ]),
    # Tier 3 — capital plan
    (3, [
        "capex", "cap ex",
        "capital expenditure", "capital plan", "capital budget",
        "hard cost", "soft cost",
        "draw schedule",
        "construction budget", "construction cost",
    ]),
    # Tier 4 — debt structure
    (4, [
        "debt", "loan", "financing", "mortgage",
    ]),
    # Tier 5 — waterfall / returns
    (5, [
        "waterfall", "promote",
        "return profile", "return metrics", "irr", "yield",
    ]),
    # Tier 6 — everything else (scan but lowest priority)
    (6, []),
]

# Short standalone sheet names (exact match against lowercased name).
# These would miss substring matching but are well-known categorical sheets.
SHEET_TIER_EXACT: dict[str, int] = {
    # Standalone cash flow / proforma sheets
    "noi":    2,
    "cf":     2,
    "cfs":    2,
    "pf":     2,   # short for "proforma"
    # CapEx
    "capex":  3,
    "cap ex": 3,
    # Debt
    "debt":   4,
    # Returns / waterfall
    "irr":    5,
}

# Sheets to SKIP entirely on first-pass extraction.
#
# Discipline:
#   - ONLY include patterns that are NEARLY CERTAIN to be noise for deal review,
#     regardless of property type
#   - When in doubt, leave the sheet in Tier 6 (scanned but low priority) rather
#     than skipping — Tier 6 will be overridden by Tier 1-5 matches anyway
#   - Skip categories should be defensible: they either DUPLICATE other data
#     (backup), or contain data that would WRONGLY match catalog metrics (comp
#     sales prices matching deal price, sensitivity IRRs matching base-case IRR)

SHEET_SKIP_PATTERNS: list[str] = [
    # Section header / navigation markers (common sponsor template convention)
    ">>>", ">>",
    # Backup / mirror data — these duplicate data from primary sheets
    "backup", "back-up", "back up",
    # Comparable / market data — would wrongly map sales-comp cap rates,
    # ADRs, or pricing to the deal's metrics
    "comp set", "compset", "comp pnl", "sales comp", "comp tab",
    # Sensitivity / scenario tables — would wrongly map alternative-case IRRs
    # and multiples to the deal's base case
    "sensitivity", "sensitivities", "sensis", "scenario",
    # Lookup / validation tables (substring matches)
    "list of values",
    "data validation", "validation tab",
]

# Short acronyms / standalone names where substring matching would either miss
# (too short to space-pad) or cause false matches. Compared exactly against the
# lowercased sheet name.
SHEET_SKIP_EXACT: set[str] = {
    # Lookup / list-of-values / reference tables
    "lov", "lovs", "lookups", "lookup",
    # Common short reference / index sheets
    "refs", "ref", "ref data", "ref tab",
}


def sheet_priority_tier(sheet_name: str) -> int:
    """
    Return the priority tier of a sheet (1=highest, 6=lowest, 99=skip).
    Tier 99 means the sheet should be skipped entirely for metric extraction.

    Resolution order:
      1. Exact-match skip set (e.g. "lov")
      2. Substring skip patterns (e.g. ">>>", "compset")
      3. Exact-match tier set (e.g. "noi" → 2)
      4. Substring tier patterns (e.g. "summary" → 1)
      5. Default Tier 6
    """
    name_lower = sheet_name.lower().strip()

    # 1. Exact-match skip
    if name_lower in SHEET_SKIP_EXACT:
        return 99

    # 2. Substring skip
    for pattern in SHEET_SKIP_PATTERNS:
        if pattern in name_lower:
            return 99

    # 3. Exact-match tier (short standalone codes)
    if name_lower in SHEET_TIER_EXACT:
        return SHEET_TIER_EXACT[name_lower]

    # 4. Substring tier patterns
    for tier, keywords in SHEET_PRIORITY_TIERS:
        for kw in keywords:
            if kw in name_lower:
                return tier

    # 5. Default
    return 6


def is_skip_sheet(sheet_name: str) -> bool:
    """True if this sheet should be skipped on first-pass extraction."""
    return sheet_priority_tier(sheet_name) == 99


def sorted_sheets_by_priority(sheet_names: list[str], exclude_skipped: bool = True) -> list[str]:
    """
    Return sheet names sorted by priority tier (highest first).
    Skip-tier sheets are excluded by default.
    """
    if exclude_skipped:
        candidates = [s for s in sheet_names if sheet_priority_tier(s) != 99]
    else:
        candidates = list(sheet_names)
    return sorted(candidates, key=lambda s: sheet_priority_tier(s))


# Column-header keywords used by table-aware extraction. Order = preference.
# When the header row of a table contains one of these keywords, the matching
# column is preferred over generic "first value to the right".
_TOTAL_COLUMN_KEYWORDS    = ["total", "total cost", "all-in", "aggregate", "lifetime", "cumulative"]
_PERIOD_COLUMN_KEYWORDS   = ["at close", "closing", "initial", "year 1", "y1", "yr 1",
                             "post close", "post-close", "stabilized", "stable", "exit", "year 5", "yr 5"]
_PER_UNIT_COLUMN_KEYWORDS = ["$/unit", "$/sf", "$/gsf", "$/nsf", "per unit", "per sf", "% total", "% of total"]


def _scan_data_columns(ws, label_row, label_col, max_scan=10):
    """Return list of column indices to the right of label_col that contain numeric values."""
    data_cols = []
    for offset in range(1, max_scan + 1):
        cell = ws.cell(row=label_row, column=label_col + offset)
        if is_numeric(cell.value):
            data_cols.append(label_col + offset)
        elif cell.value is not None and str(cell.value).strip():
            # Hit a text cell — table row ends here
            break
    return data_cols


def _detect_column_headers(ws, label_row, data_cols, max_lookback=10):
    """
    Look up from the labeled row to find the nearest header row.

    A row qualifies as a header if:
      - At least half of the data_cols have non-numeric text values in that row
      - Headers are typical column descriptors (years, periods, "Total", etc.)

    Returns {col_idx: header_text}.
    """
    if not data_cols:
        return {}

    for r_offset in range(1, max_lookback + 1):
        header_row = label_row - r_offset
        if header_row < 1:
            break
        headers = {}
        for col in data_cols:
            val = ws.cell(row=header_row, column=col).value
            if isinstance(val, str) and val.strip() and not is_numeric(val):
                headers[col] = val.strip()
        # Accept this as a header row if most data cols are text-headed
        if len(headers) >= max(2, len(data_cols) // 2):
            return headers
    return {}


def _pick_column_for_metric(headers: dict, metric_name_lower: str) -> int | None:
    """
    Given a {col: header_text} map and a metric name, pick the best column.

    Logic (in order):
      1. If metric name mentions a specific period (e.g. "Year 1", "At Close",
         "Stabilized"), use the column whose header matches that period.
      2. If metric name implies a total ("Total X", "All-in", "Project Cost"),
         use the column whose header matches Total-like keywords.
      3. Skip per-unit / per-SF / % columns (those are derived metrics, not the value).
      4. Otherwise return None — caller falls back to "first value" behavior.
    """
    headers_lower = {col: h.lower() for col, h in headers.items()}

    # (1) Period-specific preference based on metric name
    period_map = [
        (["at close", "closing", "initial", "going-in", "going in", "purchase"],
            ["at close", "closing", "initial", "going-in", "going in"]),
        (["post close", "post-close", "draws", "construction"],
            ["post close", "post-close", "draws", "construction"]),
        (["stabilized", "stabilization", "stable"],
            ["stabilized", "stable", "stab"]),
        (["year 1", "y1", "yr 1", "first year"],
            ["year 1", "y1", "yr 1"]),
        (["exit", "year 5", "yr 5", "terminal", "disposition"],
            ["exit", "year 5", "yr 5", "terminal"]),
    ]
    for metric_keywords, header_keywords in period_map:
        if any(mk in metric_name_lower for mk in metric_keywords):
            for col, h in headers_lower.items():
                if any(hk in h for hk in header_keywords):
                    return col

    # (2) Total-like preference (default for cost/proceeds/sources/uses items)
    # If ANY column header is "Total" or similar, prefer it.
    for col, h in headers_lower.items():
        if any(kw == h or kw in h for kw in _TOTAL_COLUMN_KEYWORDS):
            # But skip if it's actually a per-unit column ("$/Total" doesn't exist
            # but be safe)
            if not any(pu in h for pu in _PER_UNIT_COLUMN_KEYWORDS):
                return col

    return None


def find_nearby_value(ws, row, col, metric_name: str = "", metric_unit: str | None = None):
    """
    Find the value associated with a labeled cell.

    Strategy:
      1. Scan data columns to the right
      2. If multiple columns exist, look UP for a header row
      3. If headers found, pick the column best matching the metric's semantics
         (Total column for total metrics, period-specific for period metrics)
      4. Otherwise fall back to first non-zero value
      5. If no values right, look below
      6. Last resort: nearby grid scan

    metric_unit (Phase 1.5a) — if "date", datetime cells are accepted as values.
    """

    # Phase 1.5a — date-anchor metrics: scan for datetime cells directly.
    if metric_unit == "date":
        for offset in range(1, 8):
            v = ws.cell(row=row, column=col + offset).value
            if is_datetime(v):
                return v, cell_address(row, col + offset), "right"
        for offset in range(1, 6):
            v = ws.cell(row=row + offset, column=col).value
            if is_datetime(v):
                return v, cell_address(row + offset, col), "below"
        return None, None, None

    # Text metrics (Asset Name, Location, etc.): find the adjacent TEXT value,
    # not a numeric one. The value is typically the first non-empty text cell
    # to the right of the label. We skip cells that look like another label
    # (i.e. cells that themselves are immediately followed by their own value).
    if metric_unit == "text":
        for offset in range(1, 8):
            v = ws.cell(row=row, column=col + offset).value
            if isinstance(v, str) and v.strip() and len(v.strip()) >= 2:
                # Reject pure-numeric-looking strings (zip codes, codes)
                stripped = v.strip().replace(",", "").replace(".", "").replace("-", "")
                if stripped.isdigit():
                    continue
                return v.strip(), cell_address(row, col + offset), "right"
        # Then try below
        for offset in range(1, 4):
            v = ws.cell(row=row + offset, column=col).value
            if isinstance(v, str) and v.strip() and len(v.strip()) >= 2:
                stripped = v.strip().replace(",", "").replace(".", "").replace("-", "")
                if stripped.isdigit():
                    continue
                return v.strip(), cell_address(row + offset, col), "below"
        return None, None, None

    data_cols = _scan_data_columns(ws, row, col)

    # Table-aware path: 2+ columns of data → likely a table
    if len(data_cols) >= 2:
        headers = _detect_column_headers(ws, row, data_cols)
        if headers and metric_name:
            preferred_col = _pick_column_for_metric(headers, metric_name.lower())
            if preferred_col is not None:
                value = ws.cell(row=row, column=preferred_col).value
                if is_numeric(value):
                    return value, cell_address(row, preferred_col), "table-column"

        # Skip per-unit/percentage columns when picking fallback
        non_derived_cols = [
            c for c in data_cols
            if not (headers and any(
                pu in headers.get(c, "").lower() for pu in _PER_UNIT_COLUMN_KEYWORDS
            ))
        ] or data_cols

        right_values = [
            (ws.cell(row=row, column=c).value, cell_address(row, c))
            for c in non_derived_cols
        ]
        non_zero = [(v, a) for v, a in right_values if v != 0]
        best_val, best_addr = (non_zero[0] if non_zero else right_values[0])
        return best_val, best_addr, "right"

    # Single value to the right (no table) — return it
    if len(data_cols) == 1:
        c = data_cols[0]
        return ws.cell(row=row, column=c).value, cell_address(row, c), "right"

    # Look directly below — but STOP at the first text cell (a new row label).
    # We must not skip past an empty cell into a different metric's row.
    for offset in range(1, 4):
        value = ws.cell(row=row + offset, column=col).value
        if is_numeric(value):
            return value, cell_address(row + offset, col), "below"
        if isinstance(value, str) and value.strip():
            break  # hit another label — this metric's value is genuinely absent

    # No DIAGONAL grid scan. The old 5x5 nearby scan grabbed up/left/diagonal
    # cells — that's how "Occupancy (empty)" picked up the diagonal
    # "Discount to Replacement Cost" cell. If the value isn't immediately to the
    # right or directly below the label, treat it as NOT FOUND rather than
    # guessing a distant cell that belongs to a different metric.
    return None, None, None


# Which data_nature values are relevant per SSOT layer.
# "mixed" metrics are always included (meaningful in both projection and actual contexts).
# "underwriting" scans all three because acquisition/closing models routinely contain both
# projected values (IRR, NOI proforma, exit cap) and actual values (closing costs paid,
# loan amount drawn, actual purchase price confirmed at closing).
_LAYER_DATA_NATURE: dict[str, set] = {
    "underwriting":    {"projection", "actual", "mixed"},
    "business_plan":   {"projection", "actual", "mixed"},
    "actuals_2020":    {"actual", "mixed"},
    "actuals_2021":    {"actual", "mixed"},
    "actuals_2022":    {"actual", "mixed"},
    "actuals_2023":    {"actual", "mixed"},
    "actuals_2024":    {"actual", "mixed"},
    "actuals_2025":    {"actual", "mixed"},
    "actuals_recent":  {"actual", "mixed"},
    "rent_roll":       {"actual", "mixed"},
    "debt":            {"actual", "mixed"},
}


def filter_catalog_for_layer(catalog: list, layer: str) -> list:
    """
    Return only the metrics relevant to a given SSOT layer.

    Two filters applied:
    1. Skip calculated metrics (metric_source == "calculated") — these are
       derived after extraction, not extracted from cells.
    2. Keep only metrics whose data_nature matches the layer's expected type.
       e.g. an underwriting file should not be scanned for Current LTV or DSCR
       (those are actual/current-state metrics).
    """
    allowed_natures = _LAYER_DATA_NATURE.get(layer, {"projection", "actual", "mixed"})
    return [
        m for m in catalog
        if m.get("metric_source", "extracted") == "extracted"
        and m.get("data_nature", "mixed") in allowed_natures
    ]


# Caps for per-sheet scanning — bound worst-case work on huge files (St Regis
# has 15M cells, the catalog only cares about labels near the top of each sheet).
_MAX_ROWS_PER_SHEET = 250
_MAX_COLS_PER_SHEET = 60


def scan_workbook_for_all_metrics(file_path, catalog):
    """
    Load the workbook ONCE and scan all catalog metrics in a single pass.

    Sheet priority taxonomy (sheet_priority_tier):
      Tier 1: Summary, Deal Summary, One Pager, Executive Summary, etc.
      Tier 2: Cash flow projections (Annual CF, Proforma, P&L, NOI)
      Tier 3: CapEx / capital plan
      Tier 4: Debt structure
      Tier 5: Waterfall / returns
      Tier 6: Other analytically interesting (not skipped)
      Tier 99: SKIP — backup, comps, sensitivities, historical, market data

    When the same metric is found on multiple sheets, the LOWER-tier sheet
    (higher priority — closer to deal-level summary) wins. This stops the
    extractor from grabbing values from sensitivity tabs, comp sets, or
    historical years and labeling them as the deal's underwriting.

    Performance:
      - openpyxl read_only mode (streams cells, avoids loading whole workbook)
      - skip-tier sheets are not opened at all
      - cell iteration is bounded at 250 rows × 60 cols per sheet

    Returns {metric_id: best_match_dict_or_None} for every metric in the catalog.
    """
    # Do NOT use read_only here. read_only mode makes random cell access
    # (used by find_nearby_value to scan nearby cells) extremely slow because
    # openpyxl seeks the file on every ws.cell(row, col) call. For metric
    # extraction we need fast random access far more than fast streaming.
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return {m["metric_id"]: None for m in catalog}

    # Pre-normalize every alias once, paired with its parent metric.
    alias_index = []
    for metric in catalog:
        for alias in metric.get("aliases", []):
            alias_text = normalize_text(alias)
            if alias_text:
                alias_index.append((alias_text, metric, alias))

    matches_by_metric: dict = {m["metric_id"]: [] for m in catalog}
    file_name = Path(file_path).name

    # Use the centralized priority taxonomy — skips noise sheets entirely
    candidate_sheets = sorted_sheets_by_priority(wb.sheetnames, exclude_skipped=True)

    for sheet_name in candidate_sheets:
        sheet_tier = sheet_priority_tier(sheet_name)
        ws = wb[sheet_name]

        # Bound iteration per sheet — needed for huge models (15M cells)
        # In read_only mode, we iterate streaming so this just exits early.
        row_count = 0
        for row in ws.iter_rows(
            min_row=1, max_row=_MAX_ROWS_PER_SHEET,
            min_col=1, max_col=_MAX_COLS_PER_SHEET,
        ):
            row_count += 1
            if row_count > _MAX_ROWS_PER_SHEET:
                break
            for cell in row:
                cell_text = normalize_text(cell.value)
                if not cell_text:
                    continue

                for alias_text, metric, original_alias in alias_index:
                    if alias_text not in cell_text:
                        continue

                    # Label quality: penalise matches where the alias is a small
                    # fraction of the cell label (e.g. "noi" inside "noi to offset
                    # interest"). An exact or near-exact label match scores 1.0;
                    # a substring-in-long-label scores proportionally lower.
                    label_ratio = len(alias_text) / max(len(cell_text), 1)
                    # Also penalise if alias appears mid-word
                    idx = cell_text.find(alias_text)
                    char_before = cell_text[idx - 1] if idx > 0 else " "
                    char_after  = cell_text[idx + len(alias_text)] if idx + len(alias_text) < len(cell_text) else " "
                    mid_word = char_before.isalpha() or char_after.isalpha()
                    if mid_word:
                        continue
                    if (
                        metric.get("metric_name") == "Hold Period"
                        and alias_text in {"term", "hold"}
                        and any(x in cell_text for x in ("loan", "debt", "maturity", "amort"))
                    ):
                        continue

                    value, value_cell, direction = find_nearby_value(
                        ws, cell.row, cell.column,
                        metric_name=metric["metric_name"],
                    )
                    if value is None:
                        continue

                    # Confidence tiers
                    if direction in ("right", "below", "table-column"):
                        confidence = "exact" if label_ratio >= 0.8 else "high"
                    else:
                        confidence = "partial" if label_ratio < 0.4 else "medium"

                    matches_by_metric[metric["metric_id"]].append({
                        "metric_id":      metric["metric_id"],
                        "metric_name":    metric["metric_name"],
                        "category":       metric["category"],
                        "definition":     metric["definition"],
                        "value":          value,
                        "source_file":    file_name,
                        "sheet":          sheet_name,
                        "sheet_tier":     sheet_tier,         # NEW: drives match ranking
                        "label_cell":     cell.coordinate,
                        "value_cell":     value_cell,
                        "matched_alias":  original_alias,
                        "confidence":     confidence,
                        "label_ratio":    round(label_ratio, 2),
                        "match_method":   direction,
                    })

    try:
        wb.close()
    except Exception:
        pass

    # Best match per metric. Ranking order:
    #   1. Sheet tier (lower = higher priority — Summary beats Cash Flow beats Backup)
    #   2. Confidence tier (exact > high > medium > partial)
    #   3. Label ratio (higher = better alias match quality)
    _CONF_TIER = {"exact": 0, "high": 1, "medium": 2, "partial": 3}
    best = {}
    for metric_id, matches in matches_by_metric.items():
        if not matches:
            best[metric_id] = None
        else:
            matches.sort(key=lambda x: (
                x.get("sheet_tier", 99),
                _CONF_TIER.get(x["confidence"], 9),
                -x.get("label_ratio", 0),
            ))
            best[metric_id] = matches[0]
    return best


# =============================================================================
# Phase 1 — candidate-based extraction
# =============================================================================
#
# scan_workbook_for_candidates returns ALL matches per metric (not just the
# best one). Downstream code applies schema-aware ranking to pick the right
# candidate — or in Phase 2, an LLM resolver disambiguates when multiple
# candidates pass schema checks.
#
# Same scan loop as scan_workbook_for_all_metrics, but stores candidates
# unfiltered.

EXTRACTOR_VERSION = "phase1_5a.v1"


def scan_workbook_for_candidates(file_path, catalog, sheet_tier_map: dict | None = None):
    """
    Find ALL candidate matches per metric across the workbook.

    sheet_tier_map (Phase 2.5): optional {sheet_name: effective_tier} produced by
    content-based GPT classification. When provided, it OVERRIDES the name-based
    tier — rescuing mis-named sheets (a cash-flow proforma named "Tab3") and
    correctly skipping content-identified comps/sensitivity sheets. Sheets mapped
    to tier 99 are skipped; others are scanned in ascending tier order.

    Returns {metric_id: [list of candidate dicts]}, each candidate:
        {metric_id, metric_name, category, value, source_file, sheet, sheet_tier,
         label_cell, value_cell, matched_alias, confidence, label_ratio, match_method}
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return {m["metric_id"]: [] for m in catalog}

    alias_index = []
    for metric in catalog:
        for alias in metric.get("aliases", []):
            alias_text = normalize_text(alias)
            if alias_text:
                alias_index.append((alias_text, metric, alias))

    candidates_by_metric: dict = {m["metric_id"]: [] for m in catalog}
    file_name = Path(file_path).name

    def _tier_for(name: str) -> int:
        if sheet_tier_map and name in sheet_tier_map:
            return sheet_tier_map[name]
        return sheet_priority_tier(name)

    # Build candidate sheet list using effective tiers (skip tier-99), tier-ordered
    candidate_sheets = sorted(
        [s for s in wb.sheetnames if _tier_for(s) != 99],
        key=_tier_for,
    )

    for sheet_name in candidate_sheets:
        sheet_tier = _tier_for(sheet_name)
        ws = wb[sheet_name]

        for row in ws.iter_rows(
            min_row=1, max_row=_MAX_ROWS_PER_SHEET,
            min_col=1, max_col=_MAX_COLS_PER_SHEET,
        ):
            for cell in row:
                cell_text = normalize_text(cell.value)
                if not cell_text:
                    continue

                for alias_text, metric, original_alias in alias_index:
                    if alias_text not in cell_text:
                        continue

                    # Label quality + mid-word filtering (same logic as scan_workbook_for_all_metrics)
                    label_ratio = len(alias_text) / max(len(cell_text), 1)
                    idx = cell_text.find(alias_text)
                    char_before = cell_text[idx - 1] if idx > 0 else " "
                    char_after  = cell_text[idx + len(alias_text)] if idx + len(alias_text) < len(cell_text) else " "
                    if char_before.isalpha() or char_after.isalpha():
                        continue
                    if (
                        metric.get("metric_name") == "Hold Period"
                        and alias_text in {"term", "hold"}
                        and any(x in cell_text for x in ("loan", "debt", "maturity", "amort"))
                    ):
                        continue

                    value, value_cell, direction = find_nearby_value(
                        ws, cell.row, cell.column,
                        metric_name=metric["metric_name"],
                        metric_unit=metric.get("unit"),
                    )
                    if value is None:
                        continue

                    if direction in ("right", "below", "table-column"):
                        confidence = "exact" if label_ratio >= 0.8 else "high"
                    else:
                        confidence = "partial" if label_ratio < 0.4 else "medium"

                    candidates_by_metric[metric["metric_id"]].append({
                        "metric_id":     metric["metric_id"],
                        "metric_name":   metric["metric_name"],
                        "category":      metric["category"],
                        "value":         value,
                        "source_file":   file_name,
                        "sheet":         sheet_name,
                        "sheet_tier":    sheet_tier,
                        "label_cell":    cell.coordinate,
                        "value_cell":    value_cell,
                        "matched_alias": original_alias,
                        "confidence":    confidence,
                        "label_ratio":   round(label_ratio, 2),
                        "match_method":  direction,
                    })

    # Sort each metric's candidates by (sheet_tier, confidence_tier, -label_ratio)
    _CONF_TIER = {"exact": 0, "high": 1, "medium": 2, "partial": 3}
    for metric_id, cands in candidates_by_metric.items():
        cands.sort(key=lambda x: (
            x.get("sheet_tier", 99),
            _CONF_TIER.get(x["confidence"], 9),
            -x.get("label_ratio", 0),
        ))

    try:
        wb.close()
    except Exception:
        pass

    return candidates_by_metric


def scan_workbook_for_metric(file_path, metric):
    """
    Search one Excel workbook for one metric.
    Returns best match or None.

    NOTE: kept for backward compatibility with v1 modules. The fast path is
    scan_workbook_for_all_metrics, which avoids reloading the workbook per metric.
    """

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return None

    aliases = metric.get("aliases", [])
    matches = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                cell_text = normalize_text(cell.value)

                if not cell_text:
                    continue

                for alias in aliases:
                    alias_text = normalize_text(alias)

                    if not alias_text:
                        continue

                    if alias_text in cell_text:
                        value, value_cell, direction = find_nearby_value(
                            ws,
                            cell.row,
                            cell.column,
                            metric_name=metric["metric_name"],
                        )

                        if value is not None:
                            confidence = "high" if direction in ["right", "below"] else "medium"

                            matches.append({
                                "metric_id": metric["metric_id"],
                                "metric_name": metric["metric_name"],
                                "category": metric["category"],
                                "definition": metric["definition"],
                                "value": value,
                                "source_file": Path(file_path).name,
                                "sheet": sheet_name,
                                "label_cell": cell.coordinate,
                                "value_cell": value_cell,
                                "matched_alias": alias,
                                "confidence": confidence,
                                "match_method": direction,
                            })

    if not matches:
        return None

    # Prefer high confidence matches first
    matches = sorted(
        matches,
        key=lambda x: 0 if x["confidence"] == "high" else 1
    )

    return matches[0]


def extract_raw_labeled_pairs(file_path, max_pairs: int = 600) -> list[dict]:
    """
    Extract ALL (sheet, label, value) pairs from a workbook without any
    catalog filtering. This is the input for Pass 2 (GPT insight pass).

    Returns a list of dicts:
        {"sheet": str, "label": str, "value": numeric, "cell": str,
         "direction": "right"|"below"|"nearby", "label_len": int}

    Quality fields (used by run_raw_insight_pass to filter noise):
      direction: "right"/"below" = label directly precedes value — high signal
                 "nearby" = value found in surrounding area — lower signal
      label_len: very short labels (< 5 chars) are often headers/indices, not metrics

    Capped at max_pairs. Priority sheets (summary, assumptions, waterfall) come first.
    """
    # Do NOT use read_only — find_nearby_value does random cell access which
    # is catastrophically slow in read_only mode (every ws.cell(row,col) call
    # requires a file seek). Same fix as in scan_workbook_for_all_metrics.
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return []

    # Use the centralized priority taxonomy — skip-tier sheets are excluded entirely
    sorted_sheets = sorted_sheets_by_priority(wb.sheetnames, exclude_skipped=True)

    pairs = []
    seen_labels: set[str] = set()

    for sheet_name in sorted_sheets:
        if len(pairs) >= max_pairs:
            break
        ws = wb[sheet_name]
        tier = sheet_priority_tier(sheet_name)

        # Bound iteration per sheet — needed for huge models
        for row in ws.iter_rows(
            min_row=1, max_row=_MAX_ROWS_PER_SHEET,
            min_col=1, max_col=_MAX_COLS_PER_SHEET,
        ):
            if len(pairs) >= max_pairs:
                break
            for cell in row:
                cell_text = clean_text(cell.value)
                if not cell_text or len(cell_text) < 3:
                    continue
                if not isinstance(cell.value, str):
                    continue

                value, value_cell, direction = find_nearby_value(
                    ws, cell.row, cell.column
                )
                if value is None:
                    continue

                key = f"{sheet_name}|{normalize_text(cell_text)}"
                if key in seen_labels:
                    continue
                seen_labels.add(key)

                pairs.append({
                    "sheet":      sheet_name,
                    "sheet_tier": tier,
                    "label":      cell_text,
                    "value":      value,
                    "cell":       value_cell,
                    "direction":  direction,
                    "label_len":  len(cell_text),
                })

    try:
        wb.close()
    except Exception:
        pass

    return pairs


def extract_time_series_rows(file_path, max_rows_per_sheet: int = 25, max_total_rows: int = 80) -> list[dict]:
    """
    Find rows in the workbook that look like multi-year time series.

    A row qualifies as a time series if:
      - It has a text label in the leftmost data column
      - 3+ numeric cells follow in consecutive columns
      - A row above has text headers that look like years (2020-2035) or
        period labels (Y1, Yr 1, Year 1, Q1, Stabilized, Exit, etc.)

    Returns list of:
      {
        "sheet": str,
        "label": str,             # row label (e.g. "Net Operating Income")
        "label_cell": str,        # cell ref of label
        "headers": [str],         # column headers (years/periods)
        "values": [number],       # aligned with headers
      }

    Cap at max_rows to avoid huge payloads. Prioritises sheets named like
    cash flow projections.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return []

    # For time series we want CF/proforma sheets first; use the centralized
    # taxonomy but resort so cash flow tier (2) leads.
    # Skip-tier sheets are excluded.
    def cf_first(name: str) -> int:
        t = sheet_priority_tier(name)
        if t == 99:
            return 99
        # Boost cash flow tier to top (we want time series from CF sheets)
        if t == 2:
            return 0
        return t

    candidates = [s for s in wb.sheetnames if sheet_priority_tier(s) != 99]
    sorted_sheets = sorted(candidates, key=cf_first)

    # Labels to skip — meta/structural rows that aren't analytically interesting
    _NOISE_LABELS = {"year", "month", "period", "day", "date", "row", "n/a"}

    series = []
    series_per_sheet: dict[str, int] = {}
    period_pattern_re = re.compile(
        r"^(20\d{2}|y(ear)?\s*\d{1,2}|yr\s*\d{1,2}|q[1-4]|fy\d{2,4}|"
        r"jan(uary)?|feb(ruary)?|mar(ch)?|apr(il)?|may|jun(e)?|jul(y)?|"
        r"aug(ust)?|sep(t|tember)?|oct(ober)?|nov(ember)?|dec(ember)?|"
        r"stabili[sz]ed|exit|going.?in|at.close|post.close|trended|untrended)",
        re.IGNORECASE,
    )
    month_name_re = re.compile(
        r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\b",
        re.IGNORECASE,
    )

    def looks_like_period_header(val) -> bool:
        """Accepts strings matching the regex OR integers in year range (2000-2100)."""
        if val is None:
            return False
        # Year stored as a number
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            return 2000 <= val <= 2100
        if is_datetime(val):
            return True
        s = str(val).strip()
        if not s:
            return False
        return bool(period_pattern_re.match(s))

    def _header_year(header: str):
        import datetime as _dt
        if isinstance(header, (_dt.datetime, _dt.date)):
            return header.year
        text = str(header)
        m = re.search(r"(20\d{2}|19\d{2})", text)
        if m:
            return int(m.group(1))
        return None

    def _detect_periodicity(headers: list[str]) -> str:
        month_like = 0
        year_like = 0
        quarter_like = 0
        for h in headers:
            text = str(h)
            if (
                month_name_re.search(text)
                or is_datetime(h)
                or re.match(r"^(20\d{2}|19\d{2})[-/]\d{1,2}[-/]\d{1,2}", text.strip())
            ):
                month_like += 1
            elif re.match(r"^(q[1-4])\b", text.strip(), re.IGNORECASE):
                quarter_like += 1
            elif re.match(r"^(20\d{2}|19\d{2}|y(ear)?\s*\d+|yr\s*\d+)", text.strip(), re.IGNORECASE):
                year_like += 1
        if month_like >= 3 and month_like >= year_like:
            return "monthly"
        if quarter_like >= 3:
            return "quarterly"
        return "annual"

    def _annualize_if_monthly(headers: list[str], values: list) -> tuple[list[str], list, str | None]:
        if _detect_periodicity(headers) != "monthly":
            return [], [], None
        annual: dict[str, float] = {}
        for h, v in zip(headers, values):
            if v is None:
                continue
            year = _header_year(h)
            if year is None:
                # Fall back to sequence years when the model uses Jan/Feb...
                # without year labels. Keep the bucket explicit rather than
                # pretending it is a calendar year.
                year = "annualized_period"
            key = str(year)
            annual[key] = annual.get(key, 0.0) + float(v)
        if not annual:
            return [], [], None
        return list(annual.keys()), list(annual.values()), "sum_monthly_columns_by_year"

    for sheet_name in sorted_sheets:
        if len(series) >= max_total_rows:
            break
        ws = wb[sheet_name]
        # Pre-scan: find candidate header rows (rows where most cells are period-like)
        header_rows: list[tuple[int, dict[int, str]]] = []
        for r in range(1, min(ws.max_row, 200)):
            row_headers: dict[int, str] = {}
            period_count = 0
            for c in range(1, min(ws.max_column, 30) + 1):
                v = ws.cell(row=r, column=c).value
                if looks_like_period_header(v):
                    row_headers[c] = str(v).strip() if not isinstance(v, (int, float)) else str(int(v))
                    period_count += 1
            if period_count >= 3:
                header_rows.append((r, row_headers))

        if not header_rows:
            continue

        # For each header row, look at subsequent rows for label + values matching the header columns
        for header_row, headers in header_rows:
            header_cols = sorted(headers.keys())
            for r in range(header_row + 1, min(ws.max_row, header_row + 80) + 1):
                if len(series) >= max_total_rows:
                    break
                if series_per_sheet.get(sheet_name, 0) >= max_rows_per_sheet:
                    break

                # Look left of the first header column for a text label
                label = None
                label_cell = None
                for c in range(1, header_cols[0]):
                    v = ws.cell(row=r, column=c).value
                    if isinstance(v, str) and v.strip() and len(v.strip()) >= 3:
                        label = v.strip()
                        label_cell = cell_address(r, c)
                        break
                if not label:
                    continue

                # Skip noise labels (Year, Month, etc. — not analytically meaningful)
                if label.lower().strip(":") in _NOISE_LABELS:
                    continue

                # Collect values aligned with header columns
                values = []
                aligned_headers = []
                for c in header_cols:
                    v = ws.cell(row=r, column=c).value
                    if is_numeric(v):
                        values.append(v)
                        aligned_headers.append(headers[c])
                    else:
                        values.append(None)
                        aligned_headers.append(headers[c])

                numeric_count = sum(1 for v in values if v is not None)
                non_zero_count = sum(1 for v in values if v is not None and v != 0)

                # Skip rows where all values are zero (no signal — typically empty
                # construction draw rows in dev models)
                if non_zero_count == 0:
                    continue

                if numeric_count >= 3:
                    periodicity = _detect_periodicity(aligned_headers)
                    annual_headers, annual_values, aggregation_method = _annualize_if_monthly(
                        aligned_headers, values
                    )
                    series.append({
                        "sheet":      sheet_name,
                        "label":      label,
                        "label_cell": label_cell,
                        "headers":    aligned_headers,
                        "values":     values,
                        "periodicity": periodicity,
                        "annualized": bool(annual_headers),
                        "aggregation_method": aggregation_method,
                        "annual_headers": annual_headers,
                        "annual_values": annual_values,
                    })
                    series_per_sheet[sheet_name] = series_per_sheet.get(sheet_name, 0) + 1

    return series


def classify_file_layer(file_name):
    """
    Classify a file by its investment lifecycle layer based on its name.
    Returns one of: 'underwriting', 'business_plan', 'actuals_2021',
    'actuals_2022', 'actuals_recent', or 'unknown'.

    These names must match ssot.KNOWN_LAYERS exactly.

    Keyword groups reflect institutional RE naming conventions:
      - 'proforma' / 'pro forma' is the most common name for an UW model
      - 'BP' alone is risky (matches too much) so we anchor with word boundaries
      - financial statements: 'fs', 'financial', 'p&l', 'income statement',
        'operating statement', 't12'
    """
    name_lower = file_name.lower()

    # --- Financial Statements / actuals (check first; "2022 P&L" should NOT
    # match business plan via the year). ---
    # We pad with leading/trailing spaces so " fs " matches "FS 2022.xlsx"
    padded = f" {name_lower} "
    actuals_keywords = [
        "financial statement", "income statement", "operating statement",
        "p&l", "pl statement", "actual", "actuals",
        " fs ", "_fs_", "_fs.", " fs.", "t12", "trailing 12",
    ]
    if any(kw in padded for kw in actuals_keywords):
        for year in ("2020", "2021", "2022", "2023", "2024", "2025"):
            if year in name_lower:
                return f"actuals_{year}"
        return "actuals_recent"

    # --- Acquisition Underwriting (proforma / UW model / deal memo / closing docs) ---
    uw_keywords = [
        "acquisition", "underwriting",
        "proforma", "pro forma", "pro-forma",
        "uw model", "deal memo",
        "closing", "settlement",  # closing statement / settlement statement
        "psa", "purchase agreement",        # purchase & sale agreement
        "ic memo", "investment committee",  # IC package
    ]
    # Word-boundary check for the short token " uw" (avoid matching "answer"!)
    uw_token_match = (
        " uw" in name_lower or "_uw" in name_lower
        or name_lower.endswith(" uw") or name_lower.endswith("_uw")
    )
    if any(kw in name_lower for kw in uw_keywords) or uw_token_match:
        return "underwriting"

    # --- Business Plan (revised plan post-acquisition) ---
    bp_keywords = [
        "business plan", "budget", "forecast", "revised plan",
        "annual plan", "asset plan", "hold plan",
    ]
    if any(kw in name_lower for kw in bp_keywords):
        return "business_plan"
    # " bp " as a standalone token (so "abp_2022.xlsx" doesn't false-match)
    if " bp " in name_lower or "_bp_" in name_lower or "_bp." in name_lower or " bp." in name_lower:
        return "business_plan"

    return "unknown"


def scan_uploaded_files(upload_dir=UPLOAD_DIR):
    """
    Scan all uploaded Excel files against the metric catalog.
    Extracts each metric from EVERY file where found, tagged by source layer,
    so the analysis can compare underwriting vs business plan vs actuals.
    """

    upload_dir = Path(upload_dir)
    REPOSITORY_DIR.mkdir(exist_ok=True)

    catalog = load_metric_catalog()

    excel_files = list(upload_dir.glob("*.xlsx")) + list(upload_dir.glob("*.xlsm"))

    extracted = []
    missing = []

    for metric in catalog:
        all_matches = []

        for file_path in excel_files:
            match = scan_workbook_for_metric(file_path, metric)

            if match:
                match["source_layer"] = classify_file_layer(file_path.name)
                all_matches.append(match)

        if all_matches:
            extracted.extend(all_matches)
        else:
            missing.append({
                "metric_id": metric["metric_id"],
                "metric_name": metric["metric_name"],
                "category": metric["category"],
                "definition": metric["definition"],
                "source": metric.get("source", ""),
                "priority": metric.get("priority", "medium"),
                "aliases": metric.get("aliases", []),
                "status": "missing"
            })

    result = {
        "status": "success",
        "total_metrics": len(catalog),
        "extracted_count": len(extracted),
        "missing_count": len(missing),
        "extracted_metrics": extracted,
        "missing_metrics": missing,
    }

    with open(REPOSITORY_DIR / "flexible_extraction_result.json", "w") as f:
        json.dump(result, f, indent=2, default=str)

    pd.DataFrame(extracted).to_csv(
        REPOSITORY_DIR / "extracted_metrics_report.csv",
        index=False
    )

    pd.DataFrame(missing).to_csv(
        REPOSITORY_DIR / "missing_metrics_report.csv",
        index=False
    )

    return result


if __name__ == "__main__":
    result = scan_uploaded_files()
    print(f"Total metrics: {result['total_metrics']}")
    print(f"Extracted: {result['extracted_count']}")
    print(f"Missing: {result['missing_count']}")
    print("Saved reports to repository/")
