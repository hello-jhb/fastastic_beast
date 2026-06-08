"""
tools.py — the agent's callable tools.

Wraps deterministic Python (extraction, SSOT writes, classification) into
small, well-described functions the LLM can call via OpenAI function-calling.

Design rules:
  - Every tool returns a JSON-serializable dict.
  - Errors are returned as {"error": "..."} rather than raised. The agent
    reads the message and reacts (this is much more forgiving than exceptions).
  - Tools never call other tools internally except through composition
    (e.g. `ingest_to_ssot` calls `classify_file` and `extract_from_file`).
  - The scenario tools (`run_deal_review`, `run_perf_vs_plan`) are the only
    tools that themselves invoke an LLM; everything else is pure Python.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import ssot
from metric_catalog import load_metric_catalog
import calculations
import logging, sys
from flexible_extractor import (
    scan_workbook_for_all_metrics,
    scan_workbook_for_candidates,
    extract_raw_labeled_pairs,
    classify_file_layer,
    filter_catalog_for_layer,
    sheet_priority_tier,
    EXTRACTOR_VERSION,
)
from metric_catalog import CATALOG_VERSION
from metric_resolver import resolve_metric, make_cache_key, RESOLVER_VERSION, build_section_record
from metric_fallback import fallback_find_metric, FALLBACK_VERSION
from metric_resolver_gpt import (
    resolve_pool_with_gpt,
    run_identity_checks,
    RESOLVER_GPT_VERSION,
)
import extraction_cache
from scenarios._llm import run_raw_insight_pass, llm_available

# Logger writes to stdout so messages show up in Streamlit Cloud logs
_log = logging.getLogger("fb.tools")
if not _log.handlers:
    _h = logging.StreamHandler(sys.stdout)
    _h.setFormatter(logging.Formatter("[fb.tools] %(asctime)s %(levelname)s %(message)s"))
    _log.addHandler(_h)
    _log.setLevel(logging.INFO)


UPLOAD_DIR = Path("uploads")


# =============================================================================
# Ingestion tools — get files into SSOT
# =============================================================================

def list_uploaded_files() -> dict[str, Any]:
    """List files currently in the uploads/ directory."""
    UPLOAD_DIR.mkdir(exist_ok=True)
    files = [f.name for f in UPLOAD_DIR.iterdir() if f.is_file() and not f.name.startswith(".")]
    return {"files": sorted(files), "count": len(files)}


def classify_file(filename: str) -> dict[str, Any]:
    """
    Classify a single file by its investment lifecycle layer.
    Uses filename heuristics; reliable when files follow conventional naming
    (e.g. 'Acquisition Underwriting.xlsx', 'Financial Statement 2022.xlsx').
    """
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        return {"error": f"File not found in uploads/: {filename}"}

    layer = classify_file_layer(filename)

    return {
        "filename": filename,
        "layer": layer,
        "confidence": "high" if layer != "unknown" else "low",
    }


def extract_from_file(filename: str, layer: str | None = None) -> dict[str, Any]:
    """
    Extract metrics from a single Excel file using the metric catalog.

    If `layer` is provided, only scans metrics relevant to that SSOT layer:
      - skips calculated metrics (derived later, not extracted from cells)
      - skips metrics whose data_nature doesn't match the layer's expected type
        (e.g. won't scan for Current LTV in an underwriting file)
    """
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        return {"error": f"File not found in uploads/: {filename}"}

    if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return {"error": f"Only Excel files supported in v2. Got: {file_path.suffix}"}

    full_catalog = load_metric_catalog()
    scan_catalog = (
        filter_catalog_for_layer(full_catalog, layer) if layer
        else [m for m in full_catalog if m.get("metric_source", "extracted") == "extracted"]
    )

    matches_by_id = scan_workbook_for_all_metrics(file_path, scan_catalog)

    extracted = []
    for match in matches_by_id.values():
        if match:
            extracted.append({
                "metric_name": match["metric_name"],
                "value":       match["value"],
                "sheet":       match["sheet"],
                "value_cell":  match["value_cell"],
                "confidence":  match["confidence"],
            })

    return {
        "filename":        filename,
        "metrics":         extracted,
        "extracted_count": len(extracted),
        "catalog_size":    len(full_catalog),
        "scanned_count":   len(scan_catalog),
        "layer":           layer,
    }


def ingest_to_ssot(filename: str) -> dict[str, Any]:
    """
    Classify + extract + write to SSOT in a single operation.
    This is the tool an agent should typically call when a file is uploaded.
    """
    classification = classify_file(filename)
    if "error" in classification:
        return classification

    layer = classification["layer"]
    if layer == "unknown":
        return {
            "error": (
                f"Could not auto-classify '{filename}' from its name. "
                "Either rename it (e.g. add 'Acquisition Underwriting', "
                "'Business Plan', or 'Financial Statement 2022' to the filename) "
                "or use the manual layer-override below."
            ),
            "needs_manual_classification": True,
            "filename": filename,
        }

    return ingest_to_ssot_with_layer(filename, layer)


# =============================================================================
# Pass 2 field list — what to ask GPT to find for each layer
# =============================================================================
#
# These are the fields the SCENARIOS need that the catalog often can't capture:
#   - inferred characteristics (property_type, deal_type, strategy)
#   - derived sums (total_debt = acquisition + construction loans)
#   - context-dependent fields (capital outlay after closing)
#   - metrics the catalog might know about but with aliases not matching this
#     specific model
#
# Each entry: {name, type, hint}
#   name: stable field name the scenario template will reference
#   type: "number" | "string" — guides GPT's output format
#   hint: short note on what to look for (helps GPT find non-obvious fields)
#
# Adding a new field here makes it available to ALL scenarios. Each scenario's
# template prompt then decides which of these fields it wants to surface.

_UNDERWRITING_FIELDS_TO_FIND = [
    # ─── Characterization (inferred, not extracted) ──────────────────────
    {"name": "property_type",       "type": "string",
     "hint": "Multifamily / Office / Industrial / Retail / Hotel / Mixed-use / Conversion / Ground-up Development"},
    {"name": "deal_type",           "type": "string",
     "hint": "Acquisition / Ground-up Development / Conversion / Value-Add Renovation / Recapitalization"},
    {"name": "strategy",            "type": "string",
     "hint": "Core / Core-Plus / Value-Add / Opportunistic — infer from cap rate, occupancy, capex intensity"},
    {"name": "investment_position", "type": "string",
     "hint": "GP/Sponsor / LP / Co-GP / JV — infer from waterfall structure if present"},
    {"name": "asset_name",          "type": "string",
     "hint": "Property name (often in row 1-3 of summary sheet)"},
    {"name": "location",            "type": "string",
     "hint": "City, State or Submarket"},

    # ─── Asset basics ────────────────────────────────────────────────────
    {"name": "total_units",         "type": "number",
     "hint": "Total residential units / hotel keys / doors. Often in unit mix total row labeled 'Total / Wtd. Avg.'"},
    {"name": "total_sf",            "type": "number",
     "hint": "Total GSF / NRA / GLA — gross or rentable square footage of the asset"},

    # ─── Debt structure ──────────────────────────────────────────────────
    {"name": "total_debt",          "type": "number",
     "hint": "SUM of ALL loans (acquisition + construction + mezz). Show derivation in label_in_file."},
    {"name": "construction_loan",   "type": "number",
     "hint": "Construction loan / future funding commitment if separate from acquisition loan"},
    {"name": "loan_term_months",    "type": "number",
     "hint": "Loan term in months (e.g. '36 months I/O + 24 amortizing' → 60)"},
    {"name": "io_period_months",    "type": "number",
     "hint": "Interest-only period in months if specified"},
    {"name": "ltv",                 "type": "number",
     "hint": "Loan-to-value (debt / value). May be implied if not labeled — debt / purchase price for stabilized, or debt / stabilized value for dev."},

    # ─── Equity & waterfall ──────────────────────────────────────────────
    {"name": "lp_equity",           "type": "number",
     "hint": "LP equity contribution amount in $"},
    {"name": "gp_equity",           "type": "number",
     "hint": "GP/Sponsor equity contribution amount in $"},
    {"name": "lp_gp_split",         "type": "string",
     "hint": "Equity split as 'XX% LP / XX% GP'"},
    {"name": "preferred_return",    "type": "number",
     "hint": "LP preferred return percentage (e.g. 0.08 for 8%)"},
    {"name": "gp_promote",          "type": "string",
     "hint": "GP promote structure (e.g. '20% above 8% pref')"},

    # ─── Capital plan ────────────────────────────────────────────────────
    {"name": "capital_outlay_after_closing", "type": "number",
     "hint": "Total spend post-close: CapEx + construction draws + interest reserve + post-close soft costs"},
    {"name": "ti_lc_budget",        "type": "number",
     "hint": "Tenant improvement and leasing commission budget if present"},

    # ─── NOI bridge ──────────────────────────────────────────────────────
    {"name": "going_in_noi",        "type": "number",
     "hint": "NOI at acquisition (year 1, T12, or as-is). May be 0 for ground-up development."},
    {"name": "stabilized_noi",      "type": "number",
     "hint": "Stabilized / target NOI after lease-up or business plan completion"},
    {"name": "noi_uplift_pct",      "type": "number",
     "hint": "Percentage increase from going-in to stabilized NOI"},

    # ─── Returns the catalog might miss ──────────────────────────────────
    {"name": "cash_on_cash_year1",  "type": "number",
     "hint": "Year 1 cash-on-cash return — cash flow after debt service / equity invested"},
    {"name": "lp_irr",              "type": "number",
     "hint": "LP-specific IRR (after promote distribution) if modeled separately from deal-level IRR"},
    {"name": "lp_equity_multiple",  "type": "number",
     "hint": "LP-specific equity multiple if modeled separately"},
    {"name": "break_even_occupancy", "type": "number",
     "hint": "Occupancy required to cover debt service + operating expenses"},

    # ─── Risk context ────────────────────────────────────────────────────
    {"name": "key_risks",           "type": "string",
     "hint": "2-3 most material risks visible from the model — lease-up risk, market rent assumption risk, construction cost overrun risk, exit cap rate sensitivity, etc."},
]

_PERF_VS_PLAN_FIELDS_TO_FIND = [
    {"name": "reporting_period",    "type": "string",
     "hint": "What period does this report cover? (Year/Quarter/Month)"},
    {"name": "property_type",       "type": "string",
     "hint": "Property type — same options as underwriting"},
    {"name": "noi",                 "type": "number",
     "hint": "Actual NOI for the period"},
    {"name": "revenue",             "type": "number",
     "hint": "Effective Gross Revenue / Total Revenue for the period"},
    {"name": "expenses",            "type": "number",
     "hint": "Total Operating Expenses for the period"},
    {"name": "occupancy",           "type": "number",
     "hint": "Physical occupancy as percentage"},
    {"name": "current_loan_balance", "type": "number",
     "hint": "Outstanding loan balance at period-end"},
    {"name": "current_ltv",         "type": "number",
     "hint": "Current LTV based on most recent value"},
    {"name": "covenant_status",     "type": "string",
     "hint": "Debt covenant compliance status if mentioned"},
    {"name": "key_observations",    "type": "string",
     "hint": "Notable variances, unusual items, one-time events"},
]


def _build_fields_to_find(layer: str, found_metric_names: list[str]) -> list[dict]:
    """
    Return the list of fields Pass 2 should look for, scoped to the layer.
    Filters out fields whose name overlaps with metrics already found
    (so Pass 2 doesn't re-do deterministic work).
    """
    if layer == "underwriting":
        base = _UNDERWRITING_FIELDS_TO_FIND
    elif layer == "business_plan":
        base = _UNDERWRITING_FIELDS_TO_FIND + _PERF_VS_PLAN_FIELDS_TO_FIND
    else:
        base = _PERF_VS_PLAN_FIELDS_TO_FIND

    # Lowercase the found metric names for fuzzy overlap check
    found_lower = {n.lower() for n in found_metric_names}

    def already_found(field_name: str) -> bool:
        fn_lower = field_name.lower().replace("_", " ")
        return any(
            fn_lower in found.lower() or found.lower() in fn_lower
            for found in found_metric_names
        )

    return [f for f in base if not already_found(f["name"])]


def _inventory_sheets_by_tier(file_path: Path) -> dict[str, Any]:
    """
    Inventory the file's sheets by priority tier. Returns:
        {
          "by_tier": {1: [...], 2: [...], ..., 6: [...], 99: [...]},
          "skipped_sheets": [list of sheets the catalog scan SKIPPED],
          "low_priority_sheets": [list of Tier 6 sheets — scanned but lowest priority],
          "all_sheets": [list of every sheet name in the file],
        }

    Used so the chat agent knows which sheets exist but weren't bulk-extracted —
    it can call read_sheet/search_file on these for follow-up questions.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        by_tier: dict[int, list[str]] = {}
        for s in wb.sheetnames:
            t = sheet_priority_tier(s)
            by_tier.setdefault(t, []).append(s)
        wb.close()
        return {
            "by_tier":             by_tier,
            "skipped_sheets":      by_tier.get(99, []),
            "low_priority_sheets": by_tier.get(6, []),
            "all_sheets":          [s for sheets in by_tier.values() for s in sheets],
        }
    except Exception:
        return {"by_tier": {}, "skipped_sheets": [], "low_priority_sheets": [], "all_sheets": []}


def _run_bounded_extraction(file_path: Path, layer: str) -> tuple[dict, str, bool]:
    """
    Phase 1 — extract the 25 bounded analyst-checklist metrics with schema
    validation and ranked candidate selection.

    Returns:
      (bounded_metrics, cache_key, was_cache_hit)

      bounded_metrics: { metric_name: record_dict, ... }
      cache_key:       the versioned cache key used (for diagnostics)
      was_cache_hit:   True if the result came from cache, False if computed
    """
    from metric_catalog import load_metric_catalog

    catalog = load_metric_catalog()
    bounded = [m for m in catalog if m.get("in_bounded_list")]

    # Compute cache key from file hash + all four version constants
    file_hash = extraction_cache.file_sha256(file_path)
    cache_key = make_cache_key(
        file_hash=file_hash,
        catalog_version=CATALOG_VERSION,
        extractor_version=EXTRACTOR_VERSION,
        resolver_version=RESOLVER_VERSION,
    )

    cached = extraction_cache.load_cached(cache_key)
    if cached and isinstance(cached.get("bounded_metrics"), dict):
        _log.info(
            "BOUNDED-METRICS cache HIT for %s (key=%s, %d metrics)",
            file_path.name, cache_key[:12], len(cached["bounded_metrics"]),
        )
        return cached["bounded_metrics"], cache_key, True

    _log.info(
        "BOUNDED-METRICS cache MISS for %s (key=%s) — running extraction on %d metrics",
        file_path.name, cache_key[:12], len(bounded),
    )

    # Phase 2.5 — content-based sheet classification. Gives both a tier map
    # (for proximity fallback) AND the raw classification (for authoritative-tab
    # nomination + forbidden-source role lookup). Silently {} without API key.
    classification: dict[str, dict] = {}
    sheet_tier_map: dict[str, int] | None = None
    if llm_available():
        try:
            from sheet_classifier import classify_sheets, effective_tier
            from flexible_extractor import sheet_priority_tier
            classification = classify_sheets(file_path) or {}
            if classification:
                import openpyxl as _opx
                _wb = _opx.load_workbook(file_path, data_only=True, read_only=True)
                sheet_tier_map = {
                    name: effective_tier(name, sheet_priority_tier(name), classification)
                    for name in _wb.sheetnames
                }
                _wb.close()
                overrides = sum(
                    1 for n in sheet_tier_map
                    if sheet_tier_map[n] != sheet_priority_tier(n)
                )
                _log.info(
                    "SHEET-CLASSIFY for %s — %d sheets classified, %d tier overrides",
                    file_path.name, len(classification), overrides,
                )
        except Exception as e:
            _log.error("Sheet classification failed for %s: %s", file_path.name, e)
            classification, sheet_tier_map = {}, None

    # === PRIMARY PATH — authoritative section reader (Phase 3) ================
    # For each section (human-review order), read the authoritative tabs as
    # structured tables and extract that section's checklist. Catalog defines
    # the objective; GPT navigates structure; validation happens per-record.
    section_results: dict[str, dict] = {}   # metric_name -> extraction dict
    if classification and llm_available():
        try:
            from section_reader import read_section, SECTION_ORDER, detect_workbook_units
            from sheet_classifier import nominate_authoritative_tabs
            nominated = nominate_authoritative_tabs(classification)

            # Detect model-wide units once (e.g. BAC One Pager: "$ in '000s").
            # Applied to every section read so values report true dollars.
            summary_tabs = nominated.get("summary", []) + nominated.get("inputs", [])
            model_units_mult = detect_workbook_units(file_path, summary_tabs)
            if model_units_mult != 1:
                _log.info("UNITS for %s — model-wide ×%s detected",
                          file_path.name, f"{model_units_mult:,}")

            # group bounded metrics by section
            by_section: dict[str, list[dict]] = {}
            for m in bounded:
                by_section.setdefault(m.get("section") or "other", []).append(m)

            # Run the section reads CONCURRENTLY — each is an independent,
            # I/O-bound GPT call. 5 sections in parallel cuts wall-clock from
            # ~sum to ~max of the individual reads (~100s → ~30s).
            from concurrent.futures import ThreadPoolExecutor, as_completed
            jobs = {sec: by_section.get(sec, []) for sec in SECTION_ORDER}
            jobs = {sec: ms for sec, ms in jobs.items() if ms}
            with ThreadPoolExecutor(max_workers=len(jobs) or 1) as ex:
                futures = {
                    ex.submit(read_section, sec, ms, file_path, nominated, model_units_mult): sec
                    for sec, ms in jobs.items()
                }
                for fut in as_completed(futures):
                    sec = futures[fut]
                    try:
                        section_results.update(fut.result())
                    except Exception as e:
                        _log.error("Section %s read crashed: %s", sec, e)
        except Exception as e:
            _log.error("Section reader failed for %s: %s", file_path.name, e)
            section_results = {}

    # === FALLBACK PATH — proximity candidates (only used where section fails) =
    candidates_by_metric = scan_workbook_for_candidates(file_path, bounded, sheet_tier_map)

    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        available_sheets = list(wb.sheetnames)
        wb.close()
    except Exception:
        available_sheets = []

    # === Build one record per metric: section reader first, proximity fallback =
    # The per-metric work is independent and the GPT fallback/resolver calls are
    # I/O-bound (~10s each). On complex models 20+ metrics fall to fallback, so
    # running this loop SEQUENTIALLY cost ~4 min. Parallelize it.
    from concurrent.futures import ThreadPoolExecutor

    _counters = {"section_hits": 0, "fallback_uses": 0}

    def _build_one(metric: dict):
        record = None
        is_section = False

        # 1) Authoritative section read (validated)
        extraction = section_results.get(metric["metric_name"])
        if extraction:
            record = build_section_record(metric, extraction, classification)
            if record is not None:
                is_section = True

        # 2) Proximity fallback (only if section read missing or rejected)
        if record is None:
            cands = candidates_by_metric.get(metric["metric_id"], [])
            record = resolve_metric(metric, cands)

            if (
                record["status"] == "missing"
                and llm_available()
                and metric.get("preferred_sheets")
            ):
                fallback_cand = fallback_find_metric(metric, file_path, available_sheets)
                if fallback_cand:
                    record = resolve_metric(metric, [fallback_cand])
                    record.setdefault("validation_notes", []).insert(
                        0,
                        f"Found via GPT fallback (catalog had 0 candidates). "
                        f"Reasoning: {fallback_cand.get('fallback_reasoning', '')[:140]}",
                    )
                    record["_via_fallback"] = True

            if record["status"] == "candidate_pool" and llm_available():
                record = resolve_pool_with_gpt(record, metric, file_path)

        record["candidates"] = record.get("candidates", [])[:5]
        return metric["metric_name"], record, is_section

    bounded_metrics: dict[str, Any] = {}
    section_hits = 0
    fallback_uses = 0
    # Bounded worker count — enough to overlap the I/O without hammering the API.
    with ThreadPoolExecutor(max_workers=min(8, len(bounded) or 1)) as ex:
        for name, record, is_section in ex.map(_build_one, bounded):
            if is_section:
                section_hits += 1
            if record.pop("_via_fallback", False):
                fallback_uses += 1
            bounded_metrics[name] = record

    _log.info(
        "BOUNDED-METRICS for %s — %d via section reader, %d via GPT fallback, "
        "%d via proximity",
        file_path.name, section_hits, fallback_uses,
        len(bounded) - section_hits - fallback_uses,
    )

    # Phase 2 — identity arithmetic cross-checks (deterministic).
    # Flags inconsistencies on the implicated metric's validation_notes
    # and downgrades to "suspicious" if the discrepancy is material.
    identity_flags = run_identity_checks(bounded_metrics)
    if identity_flags:
        _log.info(
            "BOUNDED-METRICS for %s — identity checks flagged %d metric(s): %s",
            file_path.name, len(identity_flags), ", ".join(identity_flags.keys()),
        )
        for metric_name, reasons in identity_flags.items():
            rec = bounded_metrics.get(metric_name)
            if not rec:
                continue
            rec["validation_notes"] = (rec.get("validation_notes") or []) + [
                f"Identity check flagged this metric: {r}" for r in reasons
            ]
            # Only downgrade if it was verified before; never upgrade
            if rec["status"] == "verified":
                rec["status"] = "suspicious"

    # Cache the result
    extraction_cache.save_cache(
        cache_key=cache_key,
        file_name=file_path.name,
        file_hash=file_hash,
        catalog_version=CATALOG_VERSION,
        extractor_version=EXTRACTOR_VERSION,
        resolver_version=RESOLVER_VERSION,
        bounded_metrics=bounded_metrics,
    )

    # Counts for diagnostic logging
    statuses: dict[str, int] = {}
    for r in bounded_metrics.values():
        statuses[r["status"]] = statuses.get(r["status"], 0) + 1
    _log.info(
        "BOUNDED-METRICS done for %s — %s",
        file_path.name,
        ", ".join(f"{k}={v}" for k, v in sorted(statuses.items())),
    )

    return bounded_metrics, cache_key, False


# Map bounded-metric name → (pass2_field, mode).
#   mode "always":     Pass 2 inference is authoritative; use it whenever present.
#                      (Property Type can't be cell-extracted — models encode it
#                       as "Asset Type - Hotel % = 1" breakdown rows, so the
#                       catalog only ever grabs a label, never the type itself.)
#   mode "if_bad":     only backfill when the catalog value is missing/suspicious/
#                      numeric (catalog text extraction is usually fine here).
_PASS2_BACKFILL_MAP = {
    "Asset Name":    ("asset_name",    "if_bad"),
    "Property Type": ("property_type", "always"),
    "Location":      ("location",      "if_bad"),
}


def _reconcile_bounded_metrics(bounded_metrics: dict, raw_insights: dict | None) -> None:
    """
    Post-extraction reconciliation (mutates bounded_metrics in place):

    1. Backfill characterization text metrics (Asset Name, Property Type,
       Location) from Pass 2 inference when the catalog couldn't find a clean
       text value. Pass 2 reads full-sheet context and infers e.g. "Hotel"
       from "Asset Type - Hotel % = 1" — something cell-matching can't do.

    2. Derive Hold Period from Purchase Date + Exit Date when the directly-
       extracted Hold Period is missing or suspicious. There is exactly one
       true hold period; if the model didn't label it cleanly we compute it
       from the two date anchors.
    """
    found = (raw_insights or {}).get("found", {}) or {}

    # --- 1. Backfill text characterization metrics from Pass 2 ---
    for metric_name, (pass2_key, mode) in _PASS2_BACKFILL_MAP.items():
        rec = bounded_metrics.get(metric_name)
        if not rec:
            continue
        cur_val = rec.get("normalized_value")
        cur_status = rec.get("status")
        if mode == "always":
            should_backfill = True
        else:  # "if_bad"
            should_backfill = (
                cur_status in ("missing", "suspicious")
                or cur_val is None
                or isinstance(cur_val, (int, float))   # text metric holding a number = wrong
                or (isinstance(cur_val, str) and cur_val.strip().replace(".", "").isdigit())
            )
        if not should_backfill:
            continue
        p2 = found.get(pass2_key)
        if isinstance(p2, dict) and p2.get("value") not in (None, "", "—"):
            rec["raw_value"]        = p2["value"]
            rec["normalized_value"] = p2["value"]
            rec["display_value"]    = str(p2["value"])
            rec["source_sheet"]     = p2.get("sheet") or "(GPT inference)"
            rec["source_cell"]      = p2.get("label_in_file") or "—"
            rec["status"]           = "inferred"
            rec.setdefault("validation_notes", []).insert(
                0, f"Backfilled from Pass 2 characterization (key={pass2_key})."
            )

    # --- 2. Derive Hold Period from date anchors ---
    hp = bounded_metrics.get("Hold Period")
    if hp and hp.get("status") in ("missing", "suspicious"):
        pd = bounded_metrics.get("Purchase Date")
        ed = bounded_metrics.get("Exit Date")
        pdv = pd.get("normalized_value") if pd else None
        edv = ed.get("normalized_value") if ed else None
        years = _years_between(pdv, edv)
        if years is not None and 0.5 <= years <= 20:
            hp["raw_value"]        = years
            hp["normalized_value"] = round(years, 1)
            hp["display_value"]    = f"{years:.1f} years"
            hp["source_sheet"]     = "(derived)"
            hp["source_cell"]      = "Exit Date − Purchase Date"
            hp["status"]           = "inferred"
            hp.setdefault("validation_notes", []).insert(
                0,
                f"Derived from Purchase Date and Exit Date "
                f"({pdv} → {edv} = {years:.1f} yrs)."
            )


def _years_between(d1, d2):
    """Return fractional years between two date-like values, or None."""
    import datetime as _dt
    def _to_date(d):
        if isinstance(d, _dt.datetime):
            return d.date()
        if isinstance(d, _dt.date):
            return d
        if isinstance(d, str):
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S"):
                try:
                    return _dt.datetime.strptime(d[:19], fmt).date()
                except ValueError:
                    continue
        return None
    a, b = _to_date(d1), _to_date(d2)
    if a is None or b is None:
        return None
    return abs((b - a).days) / 365.25


def ingest_to_ssot_with_layer(filename: str, layer: str) -> dict[str, Any]:
    """
    Classify + extract + GPT insight pass + write to SSOT.

    Two-pass ingest:
      Pass 1 (deterministic): metric catalog extraction → structured SSOT metrics
      Pass 2 (GPT):           raw labeled-pair read → inferred characteristics,
                              gap-filled metrics, key observations

    Pass 2 runs at ingest time so every downstream scenario benefits automatically.
    It uses gpt-4o-mini to stay cheap (~$0.01–0.02 per file).
    If no API key is set, Pass 2 is silently skipped.
    """
    if layer not in ssot.KNOWN_LAYERS:
        return {"error": f"Unknown layer: {layer!r}. Valid: {sorted(ssot.KNOWN_LAYERS)}"}

    import time
    t_start = time.time()

    file_path = UPLOAD_DIR / filename

    # --- Inventory all sheets by tier so the chat agent knows what's available
    # for on-demand reads (sensitivities, scenarios, comps were intentionally
    # skipped during bulk extraction but they're still in the file).
    t0 = time.time()
    sheet_inventory = _inventory_sheets_by_tier(file_path)
    t_inventory = time.time() - t0
    _log.info(
        "INGEST %s — sheet inventory done in %.1fs (%d total, %d skipped, %d scanned)",
        filename, t_inventory,
        len(sheet_inventory["all_sheets"]),
        len(sheet_inventory["skipped_sheets"]),
        len(sheet_inventory["all_sheets"]) - len(sheet_inventory["skipped_sheets"]),
    )

    # --- Pass 1: deterministic metric extraction (legacy single-best-match path)
    t0 = time.time()
    extraction = extract_from_file(filename, layer=layer)
    t_pass1 = time.time() - t0
    if "error" in extraction:
        return extraction

    _log.info(
        "INGEST %s — Pass 1 done in %.1fs (%d metrics found, llm=%s)",
        filename, t_pass1, extraction["extracted_count"], llm_available(),
    )

    # --- Phase 1 bounded-metric extraction (schema-validated candidate ranking)
    # Runs alongside the legacy path so both old and new SSOT consumers work.
    t0 = time.time()
    try:
        bounded_metrics, bounded_cache_key, bounded_cache_hit = _run_bounded_extraction(
            file_path, layer
        )
    except Exception as e:
        _log.error("BOUNDED-METRICS failed for %s — %s: %s", filename, type(e).__name__, e)
        bounded_metrics, bounded_cache_key, bounded_cache_hit = {}, "", False
    t_bounded = time.time() - t0
    _log.info(
        "INGEST %s — bounded extraction done in %.1fs (%d metrics, cache_%s)",
        filename, t_bounded, len(bounded_metrics),
        "HIT" if bounded_cache_hit else "MISS",
    )

    # --- Pass 2: targeted GPT gap-fill + surface insights ---
    raw_insights: dict[str, Any] | None = None
    if llm_available():
        found_names = [m["metric_name"] for m in extraction["metrics"]]
        fields_to_find = _build_fields_to_find(layer, found_names)

        t0 = time.time()
        labeled_pairs = extract_raw_labeled_pairs(file_path)
        t_pairs = time.time() - t0
        _log.info(
            "INGEST %s — extracted %d raw pairs in %.1fs",
            filename, len(labeled_pairs), t_pairs,
        )

        t0 = time.time()
        raw_insights = run_raw_insight_pass(
            labeled_pairs,
            layer,
            filename,
            found_metric_names=found_names,
            fields_to_find=fields_to_find,
        )
        t_pass2 = time.time() - t0
        _log.info("INGEST %s — Pass 2 (GPT) done in %.1fs", filename, t_pass2)

    # --- Reconciliation: backfill / derive bounded metrics using Pass 2 + dates ---
    if bounded_metrics:
        _reconcile_bounded_metrics(bounded_metrics, raw_insights)

    # --- Phase 2.5 / C: comprehension verification over the fully reconciled set.
    # Holistic "does this deal cohere?" review — catches cross-metric
    # inconsistencies the deterministic identity checks miss. Flags (not silent
    # edits): downgrades clearly-contradicted metrics to suspicious + notes.
    if bounded_metrics and llm_available():
        try:
            from metric_resolver_gpt import run_comprehension_review
            review = run_comprehension_review(bounded_metrics)
            for flag in review.get("flags", []):
                mname = flag.get("metric")
                rec = bounded_metrics.get(mname)
                if not rec:
                    continue
                rec["validation_notes"] = (rec.get("validation_notes") or []) + [
                    f"Comprehension review ({flag.get('severity','?')}): {flag.get('issue','')}"
                ]
                if flag.get("severity") == "high" and rec.get("status") in ("verified", "inferred", "candidate_pool"):
                    rec["status"] = "suspicious"
            if review.get("overall"):
                _log.info("COMPREHENSION %s — %s", filename, review["overall"][:120])
        except Exception as e:
            _log.error("Comprehension review failed for %s: %s", filename, e)

    # Write both passes + bounded metrics to SSOT (sheet inventory included so
    # the agent can see which sheets exist but were intentionally not bulk-extracted)
    ssot.write_layer(
        layer=layer,
        metrics=extraction["metrics"],
        source_file=filename,
        raw_insights=raw_insights,
        sheet_inventory=sheet_inventory,
        bounded_metrics=bounded_metrics,
    )

    # Recompute derived metrics now that SSOT has new data
    calc_result = calculations.calculate_derived_metrics()

    t_total = time.time() - t_start
    _log.info("INGEST %s — TOTAL %.1fs", filename, t_total)

    # Phase 1 status breakdown for the ingest-result message
    bounded_status_counts: dict[str, int] = {}
    for r in bounded_metrics.values():
        bounded_status_counts[r["status"]] = bounded_status_counts.get(r["status"], 0) + 1

    return {
        "filename":             filename,
        "layer":                layer,
        "metric_count":         extraction["extracted_count"],
        "scanned_count":        extraction.get("scanned_count", extraction["extracted_count"]),
        "catalog_size":         extraction["catalog_size"],
        "layers_now_present":   ssot.list_layers(),
        "calculated":           calc_result["computed"],
        "insight_pass":         "completed" if raw_insights else "skipped (no API key)",
        # Sheets the agent can re-read on demand for follow-up questions
        "skipped_sheets":       sheet_inventory["skipped_sheets"],
        "low_priority_sheets":  sheet_inventory["low_priority_sheets"],
        # Phase 1 bounded-metric summary
        "bounded_metric_count":  len(bounded_metrics),
        "bounded_status_counts": bounded_status_counts,
        "bounded_cache":         "HIT" if bounded_cache_hit else "MISS",
    }


# =============================================================================
# SSOT read tools
# =============================================================================

def get_ssot_summary() -> dict[str, Any]:
    """Compact summary: layers present, files ingested, last update time."""
    return ssot.ssot_summary()


def get_layer_details(layer: str) -> dict[str, Any]:
    """Return all metrics stored in one SSOT layer, plus the sheet inventory."""
    layer_data = ssot.read_layer(layer)
    if not layer_data:
        return {"error": f"Layer '{layer}' is not present in SSOT yet."}

    inventory = layer_data.get("sheet_inventory") or {}
    return {
        "layer":         layer,
        "source_file":   layer_data["source_file"],
        "metric_count":  layer_data["metric_count"],
        "metrics":       layer_data["metrics"],
        # Sheets the catalog intentionally skipped during bulk extraction —
        # the agent should use read_sheet/search_file to inspect these when
        # the user asks about sensitivities, scenarios, comps, etc.
        "skipped_sheets":      inventory.get("skipped_sheets", []),
        "low_priority_sheets": inventory.get("low_priority_sheets", []),
        "all_sheets":          inventory.get("all_sheets", []),
    }


def check_scenario_ready(scenario: str) -> dict[str, Any]:
    """Check whether SSOT has enough data to run a given scenario."""
    return ssot.scenario_ready(scenario)


# =============================================================================
# Scenario tools — the only tools that themselves invoke an LLM
# =============================================================================

def run_deal_review() -> dict[str, Any]:
    """
    Run the Deal Review scenario. Reads the underwriting layer from SSOT and
    returns an executive summary + missing-info checklist.
    """
    from scenarios.deal_review import generate_deal_review
    return generate_deal_review()


def run_perf_vs_plan() -> dict[str, Any]:
    """
    Run the Performance vs Plan scenario. Reads UW (or BP) + actuals from SSOT
    and returns a chronological variance narrative.
    """
    from scenarios.perf_vs_plan import generate_perf_vs_plan
    return generate_perf_vs_plan()


# =============================================================================
# File inspection tools — let the agent go back to the source file on demand
# =============================================================================

def list_sheets(filename: str) -> dict[str, Any]:
    """List sheet names + dimensions for an uploaded Excel file."""
    import openpyxl
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        return {"error": f"File not found: {filename}"}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append({
                "name": name,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
            })
        wb.close()
        return {"filename": filename, "sheets": sheets, "count": len(sheets)}
    except Exception as e:
        return {"error": f"Failed to read {filename}: {type(e).__name__}: {e}"}


def read_sheet(filename: str, sheet_name: str, max_rows: int = 80) -> dict[str, Any]:
    """
    Read non-empty cells from a specific sheet. Returns up to max_rows rows
    of labeled values so the agent can read the structure on demand.
    """
    import openpyxl
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        return {"error": f"File not found: {filename}"}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # Case-insensitive sheet matching, partial-match fallback
        target = None
        for name in wb.sheetnames:
            if name.lower() == sheet_name.lower():
                target = name
                break
        if target is None:
            for name in wb.sheetnames:
                if sheet_name.lower() in name.lower():
                    target = name
                    break
        if target is None:
            return {
                "error": f"Sheet '{sheet_name}' not found in {filename}",
                "available_sheets": wb.sheetnames,
            }

        ws = wb[target]
        rows_out = []
        for r in range(1, min(ws.max_row, max_rows) + 1):
            row_cells = []
            for c in range(1, min(ws.max_column, 30) + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip() != "":
                    cell_ref = openpyxl.utils.get_column_letter(c) + str(r)
                    row_cells.append({"cell": cell_ref, "value": v})
            if row_cells:
                rows_out.append(row_cells)
        return {
            "filename":   filename,
            "sheet":      target,
            "row_count":  len(rows_out),
            "max_row":    ws.max_row,
            "max_col":    ws.max_column,
            "rows":       rows_out,
            "truncated":  ws.max_row > max_rows,
        }
    except Exception as e:
        return {"error": f"Failed to read sheet: {type(e).__name__}: {e}"}


def search_file(filename: str, query: str, max_matches: int = 30) -> dict[str, Any]:
    """
    Search every sheet for cells whose text contains the query (case-insensitive).
    Returns location + value of each match so the agent can find specific
    metrics the catalog didn't capture.
    """
    import openpyxl
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        return {"error": f"File not found: {filename}"}
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        q = query.lower().strip()
        matches = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value).lower()
                    if q in text:
                        # Also grab the nearby value (cell to the right or below)
                        nearby = None
                        right = ws.cell(row=cell.row, column=cell.column + 1).value
                        below = ws.cell(row=cell.row + 1, column=cell.column).value
                        if isinstance(right, (int, float)):
                            nearby = {"location": "right", "value": right}
                        elif isinstance(below, (int, float)):
                            nearby = {"location": "below", "value": below}
                        matches.append({
                            "sheet":  sheet_name,
                            "cell":   cell.coordinate,
                            "label":  cell.value,
                            "nearby_value": nearby,
                        })
                        if len(matches) >= max_matches:
                            break
                if len(matches) >= max_matches:
                    break
            if len(matches) >= max_matches:
                break
        wb.close()
        return {
            "filename":      filename,
            "query":         query,
            "match_count":   len(matches),
            "matches":       matches,
            "truncated":     len(matches) >= max_matches,
        }
    except Exception as e:
        return {"error": f"Search failed: {type(e).__name__}: {e}"}


# =============================================================================
# OpenAI function-calling schemas
# =============================================================================

TOOL_SCHEMAS: dict[str, dict[str, Any]] = {
    "list_uploaded_files": {
        "type": "function",
        "function": {
            "name": "list_uploaded_files",
            "description": "List files currently sitting in the uploads/ folder, so you can see what the user has provided.",
            "parameters": {"type": "object", "properties": {}, "required": []},
        },
    },
    "classify_file": {
        "type": "function",
        "function": {
            "name": "classify_file",
            "description": "Classify a single uploaded file by its investment lifecycle layer (underwriting, business_plan, actuals_2021, actuals_2022, etc.). Filename-based heuristic.",
            "parameters": {
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "Name of a file in the uploads/ folder."},
                },
                "required": ["filename"],
            },
        },
    },
    "extract_from_file": {
        "type": "function",
        "function": {
            "name": "extract_from_file",
            "description": "Run the metric catalog against one Excel file and return all metrics it finds.",
            "parameters": {
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "Name of a file in the uploads/ folder."},
                },
                "required": ["filename"],
            },
        },
    },
    "ingest_to_ssot": {
        "type": "function",
        "function": {
            "name": "ingest_to_ssot",
            "description": "Classify + extract + write to SSOT in one operation. This is the standard way to onboard a file. Call this for each uploaded file.",
            "parameters": {
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "Name of a file in the uploads/ folder."},
                },
                "required": ["filename"],
            },
        },
    },
    "get_ssot_summary": {
        "type": "function",
        "function": {
            "name": "get_ssot_summary",
            "description": "Get a compact summary of what's currently in SSOT: which layers, which files were ingested, last update time. Call this to orient yourself.",
            "parameters": {"type": "object", "properties": {}, "required": []},
        },
    },
    "get_layer_details": {
        "type": "function",
        "function": {
            "name": "get_layer_details",
            "description": "Get all metrics stored in one SSOT layer (e.g. underwriting, actuals_2022). Use this when you need specific numbers to cite.",
            "parameters": {
                "type": "object",
                "properties": {
                    "layer": {
                        "type": "string",
                        "description": "Layer name like 'underwriting', 'business_plan', 'actuals_2021', 'actuals_2022'.",
                    },
                },
                "required": ["layer"],
            },
        },
    },
    "check_scenario_ready": {
        "type": "function",
        "function": {
            "name": "check_scenario_ready",
            "description": "Check whether SSOT has enough data to run a given scenario. Returns {ready: true/false, reason, layers_present}.",
            "parameters": {
                "type": "object",
                "properties": {
                    "scenario": {
                        "type": "string",
                        "enum": ["deal_review", "perf_vs_plan"],
                    },
                },
                "required": ["scenario"],
            },
        },
    },
    "run_deal_review": {
        "type": "function",
        "function": {
            "name": "run_deal_review",
            "description": "Generate the Deal Review narrative. Call this ONLY after the underwriting layer is in SSOT. Returns markdown text summarizing the deal thesis and listing missing data.",
            "parameters": {"type": "object", "properties": {}, "required": []},
        },
    },
    "run_perf_vs_plan": {
        "type": "function",
        "function": {
            "name": "run_perf_vs_plan",
            "description": "Generate the Performance vs Plan narrative. Call this ONLY after both a plan layer (UW or BP) AND at least one actuals layer are in SSOT. Returns markdown text with chronological variance analysis.",
            "parameters": {"type": "object", "properties": {}, "required": []},
        },
    },
    "list_sheets": {
        "type": "function",
        "function": {
            "name": "list_sheets",
            "description": "List sheet names and dimensions of an uploaded Excel file. Use this when the user asks about a specific sheet by name.",
            "parameters": {
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "Name of a file in the uploads/ folder."},
                },
                "required": ["filename"],
            },
        },
    },
    "read_sheet": {
        "type": "function",
        "function": {
            "name": "read_sheet",
            "description": "Read non-empty cells from a specific sheet of an uploaded file. Use this when the user asks about content in a sheet that wasn't captured by the catalog extraction (e.g. 'what's in the Growth Rate sheet?'). Returns up to 80 rows.",
            "parameters": {
                "type": "object",
                "properties": {
                    "filename":   {"type": "string", "description": "Name of a file in uploads/."},
                    "sheet_name": {"type": "string", "description": "Sheet name (case-insensitive, partial match allowed)."},
                    "max_rows":   {"type": "integer", "description": "Max rows to return (default 80)."},
                },
                "required": ["filename", "sheet_name"],
            },
        },
    },
    "search_file": {
        "type": "function",
        "function": {
            "name": "search_file",
            "description": "Search every sheet of an uploaded file for cells containing a text query (case-insensitive). Returns matches with cell location and nearby value. Use this when the user asks about a metric or concept that isn't in the structured catalog data.",
            "parameters": {
                "type": "object",
                "properties": {
                    "filename": {"type": "string", "description": "Name of a file in uploads/."},
                    "query":    {"type": "string", "description": "Text to search for (e.g. 'rent growth', 'cap rate', 'reserve')."},
                    "max_matches": {"type": "integer", "description": "Max matches to return (default 30)."},
                },
                "required": ["filename", "query"],
            },
        },
    },
}


# Tool name -> Python implementation
TOOL_IMPLEMENTATIONS: dict[str, Any] = {
    "list_uploaded_files": list_uploaded_files,
    "classify_file": classify_file,
    "extract_from_file": extract_from_file,
    "ingest_to_ssot": ingest_to_ssot,
    "get_ssot_summary": get_ssot_summary,
    "get_layer_details": get_layer_details,
    "check_scenario_ready": check_scenario_ready,
    "run_deal_review": run_deal_review,
    "run_perf_vs_plan": run_perf_vs_plan,
    "list_sheets":  list_sheets,
    "read_sheet":   read_sheet,
    "search_file":  search_file,
}


# Tool subsets exposed per scenario. The Deal Review agent literally cannot
# call run_perf_vs_plan, and vice versa. This is what prevents v1's failure
# mode (the agent inventing scenarios that weren't asked for).
_SHARED_TOOLS = [
    "list_uploaded_files",
    "classify_file",
    "extract_from_file",
    "ingest_to_ssot",
    "get_ssot_summary",
    "get_layer_details",
    "check_scenario_ready",
    "list_sheets",   # follow-up Q&A: "what sheets are in the file?"
    "read_sheet",    # follow-up Q&A: "what's in the Growth Rate sheet?"
    "search_file",   # follow-up Q&A: "find anything mentioning rent growth"
]

TOOLS_FOR_DEAL_REVIEW = _SHARED_TOOLS + ["run_deal_review"]
TOOLS_FOR_PERF_VS_PLAN = _SHARED_TOOLS + ["run_perf_vs_plan"]


def get_tool_schemas(tool_names: list[str]) -> list[dict[str, Any]]:
    """Return the OpenAI tool-schemas list for a given subset of tool names."""
    return [TOOL_SCHEMAS[name] for name in tool_names if name in TOOL_SCHEMAS]


def call_tool(tool_name: str, arguments: dict[str, Any]) -> dict[str, Any]:
    """
    Dispatch a tool call. Used by the agent loop. Catches exceptions and
    returns them as error dicts so the agent can recover.
    """
    impl = TOOL_IMPLEMENTATIONS.get(tool_name)
    if impl is None:
        return {"error": f"Unknown tool: {tool_name}"}
    try:
        return impl(**(arguments or {}))
    except TypeError as e:
        return {"error": f"Bad arguments for {tool_name}: {e}"}
    except Exception as e:
        return {"error": f"{tool_name} crashed: {type(e).__name__}: {e}"}
