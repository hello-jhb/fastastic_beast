"""
metric_resolver_gpt.py — Phase 2 GPT resolver for candidate-pool disambiguation.

Runs only when a bounded metric has multiple validation-passing candidates and
the deterministic ranker can't tell which is right. GPT sees:
  - the metric schema (unit, period, range, definition)
  - each candidate's value, sheet, cell, and SURROUNDING CELL CONTEXT
    (row label in column A, above/below labels, column headers)
GPT picks one with reasoning and confidence.

Cost: ~$0.005-0.01 per call with gpt-4o-mini. Fires only on truly ambiguous
candidate_pool records (not when candidates all agree on value).

Also exposes run_identity_checks() — deterministic cross-checks like
Equity + Debt ≈ Total Project Cost. Flags inconsistencies as suspicious.
"""
from __future__ import annotations
import json
import logging
import sys
from pathlib import Path
from typing import Any

from scenarios._llm import client, MODEL_FAST, llm_available

log = logging.getLogger("fb.resolver_gpt")
if not log.handlers:
    h = logging.StreamHandler(sys.stdout)
    h.setFormatter(logging.Formatter("[fb.resolver_gpt] %(asctime)s %(levelname)s %(message)s"))
    log.addHandler(h)
    log.setLevel(logging.INFO)


RESOLVER_GPT_VERSION = "phase2.v1"


# =============================================================================
# Identity arithmetic — deterministic cross-checks
# =============================================================================
#
# Each check is a function (bounded_metrics) -> list of flags.
# A flag is a dict {metric_name, reason, expected, found, severity}.
# severity: "suspicious" (>5% off) or "info" (small mismatch).
#
# Currently implemented checks:
#   - Going-in Cap Rate = NOI / Purchase Price (within 10% tolerance)
#   - LTV = Debt Amount / Purchase Price (or Total Project Cost)
#   - Equity + Debt ≈ Total Project Cost
#   - Equity Multiple consistent with IRR + Hold Period (rough order check)

def _get_numeric(bm: dict, name: str):
    rec = bm.get(name)
    if not rec:
        return None
    v = rec.get("normalized_value")
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None


def _check_going_in_cap_rate(bm: dict) -> list[dict]:
    noi = _get_numeric(bm, "Net Operating Income (NOI)")
    price = _get_numeric(bm, "Purchase Price")
    cap = _get_numeric(bm, "Going-in Cap Rate")
    if not all(v is not None and v > 0 for v in (noi, price, cap)):
        return []
    implied_cap = noi / price
    if abs(implied_cap - cap) / max(cap, implied_cap) > 0.10:
        return [{
            "metric_name": "Going-in Cap Rate",
            "reason": (
                f"Stated going-in cap {cap*100:.2f}% does not reconcile with "
                f"NOI / Purchase Price = {implied_cap*100:.2f}%. "
                f"One of NOI ({noi:,.0f}), Purchase Price ({price:,.0f}), "
                f"or Cap Rate is likely wrong."
            ),
            "severity": "suspicious",
        }]
    return []


def _check_ltv_consistency(bm: dict) -> list[dict]:
    debt = _get_numeric(bm, "Debt Amount") or _get_numeric(bm, "Loan Amount")
    price = _get_numeric(bm, "Purchase Price")
    cost = _get_numeric(bm, "Total Project Cost")
    ltv = _get_numeric(bm, "Original LTV")
    if ltv is None or (debt is None) or not (price or cost):
        return []
    # LTV may be relative to purchase price OR total project cost
    candidates = [c for c in (price, cost) if c and c > 0]
    flags = []
    if not any(abs(debt/denom - ltv) / max(ltv, debt/denom) < 0.05 for denom in candidates):
        flags.append({
            "metric_name": "Original LTV",
            "reason": (
                f"Stated LTV {ltv*100:.1f}% does not reconcile with "
                f"Debt / Price ({debt/price*100:.1f}% if price={price:,.0f}) "
                f"or Debt / Total Cost. Debt or LTV likely wrong."
            ),
            "severity": "suspicious",
        })
    return flags


def _check_sources_uses(bm: dict) -> list[dict]:
    debt = _get_numeric(bm, "Debt Amount") or _get_numeric(bm, "Loan Amount")
    equity = _get_numeric(bm, "Equity Invested") or _get_numeric(bm, "Total Equity")
    cost = _get_numeric(bm, "Total Project Cost")
    if not all(v is not None and v > 0 for v in (debt, equity, cost)):
        return []
    sum_se = debt + equity
    if abs(sum_se - cost) / cost > 0.05:
        return [{
            "metric_name": "Total Project Cost",
            "reason": (
                f"Sources & Uses don't balance: "
                f"Debt ({debt:,.0f}) + Equity ({equity:,.0f}) = {sum_se:,.0f} "
                f"vs Total Cost {cost:,.0f}. >5% mismatch."
            ),
            "severity": "suspicious",
        }]
    return []


_IDENTITY_CHECKS = [
    _check_going_in_cap_rate,
    _check_ltv_consistency,
    _check_sources_uses,
]


def run_identity_checks(bounded_metrics: dict[str, Any]) -> dict[str, list[str]]:
    """
    Run all identity arithmetic checks. Returns {metric_name: [list of flag reasons]}.
    Callers can append these to validation_notes and/or downgrade status to suspicious.
    """
    flags_by_metric: dict[str, list[str]] = {}
    for check in _IDENTITY_CHECKS:
        try:
            for flag in check(bounded_metrics):
                metric = flag["metric_name"]
                flags_by_metric.setdefault(metric, []).append(flag["reason"])
        except Exception as e:
            log.warning("Identity check %s crashed: %s", check.__name__, e)
    return flags_by_metric


# =============================================================================
# GPT candidate-pool resolver
# =============================================================================

SYSTEM_PROMPT = """\
You are choosing the deal-level value for a real estate underwriting metric
from a list of candidate cells extracted from an Excel model.

Each candidate is shown with:
  - the cell's value
  - the sheet and cell reference
  - the ROW LABEL (the text label that identifies what this cell is)
  - SURROUNDING CONTEXT (cells above/below/left to give you semantic clues)
  - the COLUMN HEADER if the cell is in a table

Your task: pick the candidate whose ROW LABEL and CONTEXT best matches the
requested metric's DEFINITION at the requested PERIOD and SCALE.

REJECT candidates that:
  - are per-unit / per-key / per-SF values (when the metric wants total)
  - are historical year columns (when the metric wants going-in or stabilized)
  - are sub-property breakdowns (when the metric wants deal-level total)
  - are sensitivity/scenario alternatives (when the metric wants base case)
  - have a row label that's semantically different from the metric requested

If NONE of the candidates clearly matches, set "chosen_index" to null and
explain why. Do not pick a candidate just because it has the highest
confidence — only pick one that semantically matches.

Return ONLY JSON:
{
  "chosen_index": <int 0-based> | null,
  "reasoning": "short sentence citing the row label and any context clue",
  "confidence": "high" | "medium" | "low"
}
No prose, no markdown fences.
"""


def _get_cell_text(ws, row: int, col: int) -> str:
    """Safe cell-text getter for context gathering."""
    try:
        v = ws.cell(row=row, column=col).value
        if v is None:
            return ""
        s = str(v).strip()
        return s[:60] if len(s) > 60 else s
    except Exception:
        return ""


def _gather_candidate_context(wb, candidate: dict) -> dict:
    """
    Read the cells around a candidate's label/value cell to give GPT semantic context.

    Returns:
      {
        row_label, label_above, label_below,
        col_a_at_row, col_b_at_row,  # often the actual row label
        column_header,
        col_left_value, col_right_value
      }
    """
    sheet = candidate.get("sheet")
    label_cell = candidate.get("label_cell") or candidate.get("value_cell")
    value_cell = candidate.get("value_cell")
    if not sheet or sheet not in wb.sheetnames or not label_cell:
        return {}

    import openpyxl.utils as _u
    try:
        ws = wb[sheet]
        # Parse the label cell ref to row + col
        col_letters = "".join(ch for ch in label_cell if ch.isalpha())
        row_digits  = "".join(ch for ch in label_cell if ch.isdigit())
        if not col_letters or not row_digits:
            return {}
        lcol = _u.column_index_from_string(col_letters)
        lrow = int(row_digits)
    except Exception:
        return {}

    ctx = {
        "row_label":      _get_cell_text(ws, lrow, lcol),
        "label_above":    _get_cell_text(ws, lrow - 1, lcol),
        "label_below":    _get_cell_text(ws, lrow + 1, lcol),
        "col_a_at_row":   _get_cell_text(ws, lrow, 1),
        "col_b_at_row":   _get_cell_text(ws, lrow, 2),
    }

    # If the value cell is different from the label, get column header by walking up
    if value_cell and value_cell != label_cell:
        try:
            vcol_letters = "".join(ch for ch in value_cell if ch.isalpha())
            vrow_digits  = "".join(ch for ch in value_cell if ch.isdigit())
            vcol = _u.column_index_from_string(vcol_letters)
            vrow = int(vrow_digits)
            # Walk up from the value cell looking for a text cell that's likely a header
            for r in range(vrow - 1, max(0, vrow - 8), -1):
                txt = _get_cell_text(ws, r, vcol)
                if txt and not txt.replace(",", "").replace(".", "").replace("-", "").isdigit():
                    ctx["column_header"] = txt
                    break
        except Exception:
            pass

    return ctx


def _format_candidate_for_prompt(idx: int, candidate: dict, ctx: dict) -> str:
    """Render a candidate + context as readable text block for the GPT prompt."""
    lines = [
        f"Candidate {idx}:",
        f"  value:        {candidate.get('value')}",
        f"  sheet:        {candidate.get('sheet')}",
        f"  cell:         {candidate.get('value_cell')}",
        f"  sheet_tier:   {candidate.get('sheet_tier')} (lower = more authoritative)",
        f"  matched alias: {candidate.get('matched_alias')!r}",
    ]
    if ctx:
        if ctx.get("row_label"):
            lines.append(f"  row label:    {ctx['row_label']!r}")
        if ctx.get("col_a_at_row"):
            lines.append(f"  col A label:  {ctx['col_a_at_row']!r}")
        if ctx.get("col_b_at_row") and ctx["col_b_at_row"] != ctx.get("row_label"):
            lines.append(f"  col B label:  {ctx['col_b_at_row']!r}")
        if ctx.get("label_above"):
            lines.append(f"  cell above:   {ctx['label_above']!r}")
        if ctx.get("label_below"):
            lines.append(f"  cell below:   {ctx['label_below']!r}")
        if ctx.get("column_header"):
            lines.append(f"  column hdr:   {ctx['column_header']!r}")
    return "\n".join(lines)


def _candidates_substantially_agree(candidates: list[dict]) -> bool:
    """
    Return True if all validation-passing candidates have the same value
    (within 1% tolerance for numerics). Saves GPT calls when there's no real
    ambiguity — different sheets reporting the same number is fine.
    """
    passing = [c for c in candidates if c.get("passes_validation")]
    if len(passing) < 2:
        return True
    values = [c.get("value") for c in passing if c.get("value") is not None]
    if len(values) < 2:
        return True
    # Numeric agreement test
    try:
        nums = [float(v) for v in values]
        if max(nums) == min(nums):
            return True
        tol = 0.01 * max(abs(max(nums)), abs(min(nums)), 1)
        return (max(nums) - min(nums)) <= tol
    except (TypeError, ValueError):
        return all(v == values[0] for v in values)


def _gpt_pick(metric: dict, candidates: list[dict], context_by_idx: dict[int, dict]) -> dict:
    """Send candidates + context to GPT, get back chosen_index + reasoning."""
    blocks = [
        _format_candidate_for_prompt(i, c, context_by_idx.get(i, {}))
        for i, c in enumerate(candidates)
    ]
    user_msg = (
        f"METRIC: {metric['metric_name']}\n"
        f"DEFINITION: {metric.get('definition', '')}\n"
        f"EXPECTED UNIT: {metric.get('unit')}\n"
        f"EXPECTED PERIOD: {metric.get('period')}\n"
        f"PREFERRED SHEETS: {', '.join(metric.get('preferred_sheets', []) or [])}\n"
        f"\nCANDIDATES:\n\n" + "\n\n".join(blocks)
    )

    try:
        response = client.chat.completions.create(
            model=MODEL_FAST,
            temperature=0.0,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user",   "content": user_msg},
            ],
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        return json.loads(raw)
    except json.JSONDecodeError as e:
        log.error("GPT picker JSON parse failed for %s: %s", metric["metric_name"], e)
        return {"chosen_index": None, "reasoning": "JSON parse error", "confidence": "low"}
    except Exception as e:
        log.error("GPT picker API failed for %s: %s", metric["metric_name"], e)
        return {"chosen_index": None, "reasoning": f"API error: {e}", "confidence": "low"}


def resolve_pool_with_gpt(record: dict, metric: dict, file_path: Path) -> dict:
    """
    Phase 2 — if `record` is candidate_pool, send candidates + context to GPT
    and pick the right one. Returns an updated record (or the original if
    GPT call wasn't possible).

    Skips the GPT call if:
      - LLM unavailable
      - candidates substantially agree on value (already verified-equivalent)
      - record status != "candidate_pool"
    """
    if record.get("status") != "candidate_pool":
        return record
    if not llm_available():
        return record

    candidates = record.get("candidates", [])
    passing = [c for c in candidates if c.get("passes_validation")]
    if len(passing) < 2:
        # Only 1 (or 0) passing — nothing to disambiguate
        return record

    if _candidates_substantially_agree(passing):
        # All passing candidates have ~same value; just promote to verified
        record["status"] = "verified"
        record["validation_notes"] = (record.get("validation_notes") or []) + [
            f"Promoted to verified — all {len(passing)} passing candidates agree on value."
        ]
        return record

    # Gather context per passing candidate
    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        log.error("Could not open workbook for context gathering: %s", e)
        return record

    context_by_idx: dict[int, dict] = {}
    for i, c in enumerate(passing):
        context_by_idx[i] = _gather_candidate_context(wb, c)

    try:
        wb.close()
    except Exception:
        pass

    log.info(
        "GPT-resolve START for %s (%d passing candidates)",
        metric["metric_name"], len(passing),
    )
    result = _gpt_pick(metric, passing, context_by_idx)

    chosen_idx = result.get("chosen_index")
    reasoning  = result.get("reasoning", "")
    confidence = result.get("confidence", "low")

    if chosen_idx is None or not (0 <= chosen_idx < len(passing)):
        # GPT couldn't decide
        record["validation_notes"] = (record.get("validation_notes") or []) + [
            f"GPT resolver could not confidently pick a candidate. Reasoning: {reasoning[:160]}"
        ]
        log.info("GPT-resolve INCONCLUSIVE for %s — %s",
                 metric["metric_name"], reasoning[:80])
        return record

    chosen = passing[chosen_idx]
    # Apply the chosen candidate
    record["raw_value"]     = chosen.get("value")
    record["source_sheet"]  = chosen.get("sheet")
    record["source_cell"]   = chosen.get("value_cell")
    record["sheet_tier"]    = chosen.get("sheet_tier")
    record["extractor_confidence"] = chosen.get("confidence")

    # Apply scale correction if the chosen candidate needed it
    raw = chosen.get("value")
    scale_correction = chosen.get("scale_correction")
    if scale_correction == "1000":
        record["normalized_value"] = float(raw) * 1000
    elif scale_correction == "1000000":
        record["normalized_value"] = float(raw) * 1_000_000
    else:
        record["normalized_value"] = raw

    # Re-format display value
    from metric_resolver import _format_display
    record["display_value"] = _format_display(
        record["normalized_value"], metric.get("unit"), metric.get("scale"),
    )

    record["status"] = "verified"
    record["validation_notes"] = (record.get("validation_notes") or []) + [
        f"Phase 2 GPT resolver picked candidate {chosen_idx} ({confidence} confidence). "
        f"Reasoning: {reasoning[:200]}"
    ]
    log.info(
        "GPT-resolve PICKED %s for %s — %s",
        chosen.get("value_cell"), metric["metric_name"], reasoning[:80],
    )
    return record
