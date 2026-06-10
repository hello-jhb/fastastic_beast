"""
section_reader.py — Phase 3 authoritative-tab section reader (the new spine).

Replaces proximity cell-hunting as the PRIMARY extraction path. Instead of
"find a number near a label," it works like a human analyst:

    1. The workbook mapper says which tabs are authoritative for each role.
    2. For a section (property / deal_basis / leverage / returns / capex), the
       catalog provides the exact metric checklist + schema + source hierarchy.
    3. The authoritative tab(s) for that section are rendered as STRUCTURED
       TABLES (row labels + column headers + cell refs) — what a human sees.
    4. GPT is asked to find ONLY the checklist metrics in that table, returning
       value + cell + column + reasoning per metric. It navigates structure;
       it does not free-form "extract everything."
    5. Every returned value is validated against the catalog (unit / range /
       period / forbidden-source) before it can enter SSOT.

This is HYBRID: catalog defines the objective and validates; GPT only reads
structure. Determinism is preserved by (a) a fixed per-section checklist,
(b) catalog validation of every value, (c) file-hash caching upstream.

Proximity (scan_workbook_for_candidates) remains available as a FALLBACK for
metrics the section reader cannot find.
"""
from __future__ import annotations
import json
import logging
import sys
from pathlib import Path
from typing import Any

import openpyxl
import openpyxl.utils as xlutils

from scenarios._llm import client, llm_available
from re_knowledge import knowledge_block

log = logging.getLogger("fb.section_reader")
if not log.handlers:
    h = logging.StreamHandler(sys.stdout)
    h.setFormatter(logging.Formatter("[fb.section_reader] %(asctime)s %(levelname)s %(message)s"))
    log.addHandler(h)
    log.setLevel(logging.INFO)

SECTION_READER_VERSION = "v1"

# gpt-4o for the structured-table read — institutional tables are where the
# stronger model earns its cost. One call per section keeps the prompt focused.
SECTION_MODEL = "gpt-4o"

# Rendering caps. Cash-flow / returns / exit tables can span 10-12 years of
# columns plus stabilized/exit, so the column cap must be generous or the
# right period column gets truncated (this caused returns 0/3 on a test file).
_MAX_ROWS = 140
_MAX_COLS = 40

# Human-review order. The orchestrator extracts sections in this sequence.
SECTION_ORDER = ["property", "deal_basis", "leverage", "returns", "capex"]


# ---------------------------------------------------------------------------
# Tab rendering — turn a sheet into a structured table a human/GPT can read
# ---------------------------------------------------------------------------

def _detect_units_multiplier(ws, max_row: int, max_col: int) -> int:
    """
    Scan the top rows of a sheet for a units declaration and return the
    multiplier to convert displayed values to true dollars:
        "in $000s" / "in thousands" / "($ in 000s)"      → 1_000
        "$MM" / "in millions" / "($ in millions)"        → 1_000_000
    Returns 1 if no marker found. Only scans the top ~25 rows (markers are
    almost always near the title).
    """
    import re as _re
    thousands_re = _re.compile(r"\$?\s*0{3}s|in\s+thousands|\(\s*\$?\s*in\s+000", _re.IGNORECASE)
    # '000s, $000s, "in thousands", "($ in 000s)"
    millions_re  = _re.compile(r"\$mm\b|in\s+millions|\(\s*\$?\s*in\s+millions|\$\s*millions", _re.IGNORECASE)
    for r in range(1, min(max_row, 25) + 1):
        for c in range(1, min(max_col, 12) + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            s = v.strip()
            if millions_re.search(s):
                return 1_000_000
            if thousands_re.search(s):
                return 1_000
    return 1


def detect_workbook_units(file_path: Path, summary_tabs: list[str]) -> int:
    """
    Detect a MODEL-WIDE units multiplier by scanning the summary/one-pager tabs.
    Institutional models often state "$ Amounts in '000s" once on the summary
    and apply it to every tab (BAC: One Pager says '000s; the Capex tab has no
    marker but its values are still in thousands). Returns 1 / 1_000 / 1_000_000.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception:
        return 1
    mult = 1
    for tab in summary_tabs:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        m = _detect_units_multiplier(ws, min(ws.max_row, 25), min(ws.max_column, 12))
        if m != 1:
            mult = m
            break
    try:
        wb.close()
    except Exception:
        pass
    return mult


def render_tab_as_table(file_path: Path, sheet_name: str, default_units_mult: int = 1) -> str:
    """
    Render one sheet as a compact, cell-referenced table.

    default_units_mult: model-wide units multiplier to apply when this specific
    tab has no units marker of its own (units are often declared once on the
    summary and apply to every tab).

    Each non-empty row is emitted as:
        <row label>  |  <colHeader>=<value> (CELL)  |  ...
    Header rows (mostly text across columns near the top) are emitted first so
    GPT can see the period columns (At Close / Year 1 / Stabilized / Exit).
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return f"(could not load workbook: {e})"
    if sheet_name not in wb.sheetnames:
        wb.close()
        return f"(sheet {sheet_name!r} not found)"

    ws = wb[sheet_name]
    max_row = min(ws.max_row, _MAX_ROWS)
    max_col = min(ws.max_column, _MAX_COLS)

    # Detect a tab-level UNITS marker ("in $000s", "in thousands", "$MM",
    # "in millions"). Hotel/large models often state this once at the top and
    # then list values in those units (e.g. BAC Capex: Total Capex = 23,940.6
    # meaning $23.94M). We surface the multiplier to GPT so it reports the true
    # dollar value, not the displayed-in-thousands figure.
    # Tab-specific marker wins; otherwise inherit the model-wide default.
    units_mult = _detect_units_multiplier(ws, max_row, max_col)
    if units_mult == 1 and default_units_mult != 1:
        units_mult = default_units_mult

    # Detect PERIOD header rows only — rows where multiple cells look like
    # period/column labels (years, Year N, Stabilized, Exit, At Close, Going-In).
    # A key-value sheet (General Information) has none → renders as pure
    # row-label: value(cell), which is correct. A matrix (cash flow) has a year
    # header row → its columns get period labels so GPT knows which period.
    import re as _re
    _period_re = _re.compile(
        r"^(20\d{2}|19\d{2}|year\s*\d+|yr\s*\d+|y\d+|q[1-4]|"
        r"stabili[sz]ed|exit|going.?in|at.?close|post.?close|trended|untrended|"
        r"t-?12|trailing|forward|reversion|terminal|"
        # column headers of sources/uses & basis build-up tables — so GPT can
        # pick the TOTAL column (deal-level) over the At-Close column.
        r"total|aggregate|all.?in|combined|consolidated|cumulative|"
        r"\$/unit|\$/sf|\$/key|\$/gsf|per unit|per sf|per key|% total)",
        _re.IGNORECASE,
    )

    def _is_period(v) -> bool:
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return 1990 <= v <= 2100  # bare year
        if isinstance(v, str) and v.strip():
            return bool(_period_re.match(v.strip()))
        return False

    period_header_rows: dict[int, dict[int, str]] = {}
    for r in range(1, min(max_row, 20) + 1):
        period_cells = {}
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if _is_period(v):
                period_cells[c] = (str(int(v)) if isinstance(v, (int, float)) else str(v).strip())[:18]
        if len(period_cells) >= 2:
            period_header_rows[r] = period_cells

    def header_for_col(col: int) -> str:
        best = ""
        for hr in sorted(period_header_rows.keys()):
            if col in period_header_rows[hr]:
                best = period_header_rows[hr][col]
        return best

    header_rows = period_header_rows  # only skip genuine period-header rows below

    lines: list[str] = [f"=== SHEET: {sheet_name} ==="]
    if units_mult and units_mult != 1:
        unit_word = "thousands ($000s)" if units_mult == 1_000 else "millions ($MM)"
        lines.append(
            f"!! UNITS NOTE: dollar values on this sheet are stated in {unit_word}. "
            f"Multiply by {units_mult:,} to get the true dollar amount "
            f"(e.g. a displayed 23,940 means ${23940*units_mult:,.0f})."
        )
    if header_rows:
        hdr_desc = []
        for hr, cells in sorted(header_rows.items()):
            cols = ", ".join(f"{xlutils.get_column_letter(c)}={t}" for c, t in sorted(cells.items()))
            hdr_desc.append(f"  header row {hr}: {cols}")
        lines.append("COLUMN HEADERS:")
        lines.extend(hdr_desc)
        lines.append("ROWS:")

    for r in range(1, max_row + 1):
        if r in header_rows:
            continue
        # Find the row label: first non-empty text cell in cols A-C
        label = ""
        for c in range(1, min(4, max_col + 1)):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                label = v.strip()[:48]
                break
        # Collect numeric/date/text values across the row
        cells = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            # Skip the label cell itself
            if isinstance(v, str) and v.strip()[:48] == label and c < 4:
                continue
            ref = xlutils.get_column_letter(c) + str(r)
            col_hdr = header_for_col(c)
            val_str = _fmt_cell(v)
            if col_hdr:
                cells.append(f"{col_hdr}={val_str}({ref})")
            else:
                cells.append(f"{val_str}({ref})")
        if not label and not cells:
            continue
        if label and not cells:
            continue  # label-only row, skip
        lines.append(f"  {label or '(no label)'}: " + " | ".join(cells[:_MAX_COLS]))

    wb.close()
    return "\n".join(lines)


def _fmt_cell(v) -> str:
    import datetime as _dt
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d") if isinstance(v, _dt.datetime) else v.isoformat()
    if isinstance(v, float):
        if abs(v) >= 1000:
            return f"{v:,.0f}"
        return f"{v:,.4g}"
    return str(v)[:40]


# ---------------------------------------------------------------------------
# Authoritative-tab resolution for a section
# ---------------------------------------------------------------------------

def _tabs_for_section(
    section_metrics: list[dict],
    nominated: dict[str, list[str]],
    max_tabs: int = 4,
) -> list[str]:
    """
    Union the primary source roles of all metrics in the section, resolve each
    role to its nominated authoritative tab(s), and return an ordered, de-duped
    list of actual sheet names to read for this section.

    If the section's preferred roles resolve to NO tabs (classifier gap on a
    complex model), fall back to the catch-all authoritative pool so the
    section still gets read rather than going silently missing.
    """
    ordered_roles: list[str] = []
    for m in section_metrics:
        for role in (m.get("source_primary") or []):
            if role not in ordered_roles:
                ordered_roles.append(role)

    tabs: list[str] = []
    for role in ordered_roles:
        for tab in nominated.get(role, []):
            if tab not in tabs:
                tabs.append(tab)

    # Catch-all fallback: ensure summary-class tabs are always available so a
    # classifier gap doesn't leave the whole section blind.
    for tab in nominated.get("_catch_all", []):
        if tab not in tabs:
            tabs.append(tab)

    return tabs[:max_tabs]


# ---------------------------------------------------------------------------
# The section read — hybrid GPT call
# ---------------------------------------------------------------------------

SECTION_SYSTEM = """\
You are a real estate analyst reading authoritative tabs of an underwriting
model to extract a SPECIFIC checklist of metrics. You are shown the tab(s) as
structured tables (row labels + column headers + cell references like C11).

""" + knowledge_block(include=["period", "deal_level", "debt", "property_type"]) + """

RULES:
- Find ONLY the metrics on the checklist. Do not report anything else.
- For each metric, choose the cell whose ROW LABEL and COLUMN HEADER match the
  metric's definition AND the requested PERIOD. The column header tells you the
  period (At Close / Year 1 / Going-In / Stabilized / Exit).
- Return the cell reference exactly as shown (e.g. "C11").
- If a metric is genuinely not present in these tabs, return found=false for it.
  Do NOT guess or substitute a nearby unrelated number.
- For text metrics (name, type, location), return the text value.
- For a metric appearing in MULTIPLE tabs with DIFFERENT values, report the one
  from the most authoritative tab and note the other in "alt".

Return ONLY JSON:
{
  "results": [
    {"metric": "<exact checklist name>",
     "found": true,
     "value": <number|string>,
     "cell": "C11",
     "sheet": "<sheet name>",
     "period": "<observed period, e.g. Year 1 / Stabilized / Exit / ''>",
     "column_header": "<the period/column label, or '' >",
     "reasoning": "<row label + column header that justify this>",
     "alt": {"value": <other value>, "cell": "...", "sheet": "..."}  // optional
    },
    {"metric": "<name>", "found": false, "reasoning": "not present in these tabs"}
  ]
}
No prose, no code fences.
"""


def read_section(
    section: str,
    section_metrics: list[dict],
    file_path: Path,
    nominated_tabs: dict[str, list[str]],
    model_units_mult: int = 1,
) -> dict[str, dict]:
    """
    Read one section from its authoritative tabs.

    model_units_mult: workbook-wide units multiplier (1 / 1_000 / 1_000_000),
    applied to tabs that don't declare their own units. GPT is told to report
    true dollar values.

    Returns {metric_name: extraction_dict} where extraction_dict is:
        {found, value, cell, sheet, column_header, reasoning, alt?}
    Only metrics GPT addressed are included. Validation happens downstream.

    Returns {} if LLM unavailable or no authoritative tabs resolved.
    """
    if not llm_available():
        return {}

    tabs = _tabs_for_section(section_metrics, nominated_tabs)
    if not tabs:
        log.info("Section %s — no authoritative tabs resolved; skipping reader", section)
        return {}

    # Render the authoritative tabs (with model-wide units applied)
    rendered = []
    for tab in tabs:
        rendered.append(render_tab_as_table(file_path, tab, default_units_mult=model_units_mult))
    tables_block = "\n\n".join(rendered)

    # Build the checklist
    checklist_lines = []
    for m in section_metrics:
        checklist_lines.append(
            f"  - {m['metric_name']}"
            f"  [unit={m.get('unit')}, period={m.get('period')}]"
            f"  — {m.get('definition', '')[:80]}"
        )
    checklist = "\n".join(checklist_lines)

    user_msg = (
        f"SECTION: {section}\n\n"
        f"CHECKLIST (find exactly these, nothing else):\n{checklist}\n\n"
        f"AUTHORITATIVE TABS:\n\n{tables_block}"
    )

    try:
        response = client.chat.completions.create(
            model=SECTION_MODEL,
            temperature=0.0,
            messages=[
                {"role": "system", "content": SECTION_SYSTEM},
                {"role": "user",   "content": user_msg},
            ],
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        parsed = json.loads(raw)
    except json.JSONDecodeError as e:
        log.error("Section %s JSON parse failed: %s", section, e)
        return {}
    except Exception as e:
        log.error("Section %s read failed: %s", section, e)
        return {}

    out: dict[str, dict] = {}
    for r in parsed.get("results", []):
        name = r.get("metric")
        if name:
            out[name] = r

    found_n = sum(1 for r in out.values() if r.get("found"))
    log.info(
        "Section %s read from tabs %s — %d/%d found",
        section, tabs, found_n, len(section_metrics),
    )
    return out
