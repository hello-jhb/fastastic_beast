"""
sheet_classifier.py — content-based sheet role classification (Phase 2.5 / B).

The deterministic sheet_priority_tier() in flexible_extractor classifies sheets
by NAME keywords. That's a weak signal — a cash-flow proforma named "Tab3" or a
sponsor's idiosyncratic "Working" sheet falls to tier 6 or gets mis-handled.

This module classifies each sheet by its CONTENT (name + first rows of labels)
using a single batched GPT call, returning a role per sheet. The role then
OVERRIDES the name-based tier when GPT is confident — so extraction becomes
robust to naming conventions we've never seen.

Runs once per file at ingest. ~$0.005. Result is cached as part of the bounded
extraction cache. Silently no-ops without an API key (falls back to name-based
tiers only).
"""
from __future__ import annotations
import json
import logging
import sys
from pathlib import Path

from scenarios._llm import client, MODEL_FAST, llm_available
from re_knowledge import SHEET_ROLE_VOCAB, ROLE_TO_TIER
from knowledge_store import build_runtime_knowledge_block

log = logging.getLogger("fb.sheet_classifier")
if not log.handlers:
    h = logging.StreamHandler(sys.stdout)
    h.setFormatter(logging.Formatter("[fb.sheet_classifier] %(asctime)s %(levelname)s %(message)s"))
    log.addHandler(h)
    log.setLevel(logging.INFO)

CLASSIFIER_VERSION = "v1"

_MAX_LABELS_PER_SHEET = 18
_MAX_SHEETS = 120


SYSTEM_PROMPT = f"""\
You classify the sheets of a real estate underwriting Excel model by ROLE,
based on the row labels you are shown — NOT the sheet name (names are
unreliable). A sheet named "Tab3" may be a cash flow proforma; a sheet named
"Summary" may be empty.

{SHEET_ROLE_VOCAB}

{build_runtime_knowledge_block(["workbook_mapping"])}

For each sheet, output its role and a confidence (high/medium/low).
Confidence is "low" if the labels are too sparse to tell.

Return ONLY JSON:
{{
  "sheets": [
    {{"name": "<exact sheet name>", "role": "<role>", "confidence": "high|medium|low"}},
    ...
  ]
}}
No prose, no code fences. Include every sheet you were given.
"""


def _gather_sheet_labels(file_path: Path) -> dict[str, list[str]]:
    """For each sheet, collect up to N text labels from columns A-C of the top rows."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception:
        return {}

    out: dict[str, list[str]] = {}
    for sheet_name in wb.sheetnames[:_MAX_SHEETS]:
        ws = wb[sheet_name]
        labels: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=6):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.strip() and len(v.strip()) >= 3:
                    s = v.strip()
                    labels.append(s[:50] if len(s) > 50 else s)
                    if len(labels) >= _MAX_LABELS_PER_SHEET:
                        break
            if len(labels) >= _MAX_LABELS_PER_SHEET:
                break
        out[sheet_name] = labels
    try:
        wb.close()
    except Exception:
        pass
    return out


def classify_sheets(file_path: Path) -> dict[str, dict]:
    """
    Classify every sheet by content. Returns:
        { sheet_name: {"role": str, "confidence": str, "implied_tier": int} }

    Returns {} if LLM unavailable (caller falls back to name-based tiers).
    """
    if not llm_available():
        return {}

    labels_by_sheet = _gather_sheet_labels(file_path)
    if not labels_by_sheet:
        return {}

    # Build the prompt: sheet name + its labels
    blocks = []
    for name, labels in labels_by_sheet.items():
        label_str = " | ".join(labels) if labels else "(no text labels found)"
        blocks.append(f'Sheet "{name}": {label_str}')
    user_msg = "Classify these sheets:\n\n" + "\n".join(blocks)

    try:
        response = client.chat.completions.create(
            model=MODEL_FAST,
            temperature=0.0,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user",   "content": user_msg},
            ],
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        parsed = json.loads(raw)
    except json.JSONDecodeError as e:
        log.error("Sheet classification JSON parse failed: %s", e)
        return {}
    except Exception as e:
        log.error("Sheet classification API failed: %s", e)
        return {}

    result: dict[str, dict] = {}
    for entry in parsed.get("sheets", []):
        name = entry.get("name")
        role = entry.get("role", "other")
        conf = entry.get("confidence", "low")
        if name is None:
            continue
        result[name] = {
            "role":         role,
            "confidence":   conf,
            "implied_tier": ROLE_TO_TIER.get(role, 8),
        }

    log.info(
        "Sheet classification done for %s — %d sheets classified",
        file_path.name, len(result),
    )
    return result


def nominate_authoritative_tabs(
    classification: dict[str, dict],
) -> dict[str, list[str]]:
    """
    From the per-sheet classification, nominate the authoritative tab(s) per role.

    Returns {role: [sheet_name, ...]} ordered best-first (high confidence first).
    Only includes non-skipped roles. Used by the Section Reader to know which
    actual tabs in THIS file to read for each metric's source hierarchy.

    Example return:
      {
        "summary":      ["One Pager", "Key UW Metrics", "General Information"],
        "inputs":       ["Inputs"],
        "cash_flow":    ["Annual CFs"],
        "debt":         ["Debt Information", "Debt"],
        "returns":      ["IRR Tracker"],
        "sources_uses": ["Inputs"],   # if a sheet is classified sources_uses
      }
    """
    # Authority is decided by CONTENT role, never by sheet name. A tab named
    # "Charts" or "Summary Charts" is often the model's dashboard one-pager with
    # the key deal outputs — if the content classifier roled it 'summary', it IS
    # authoritative and must be read. Junk is excluded by ROLE (comps /
    # sensitivity / backup / other), which the classifier assigns from content.
    _CONF_RANK = {"high": 0, "medium": 1, "low": 2}
    by_role: dict[str, list[tuple[int, str]]] = {}
    for sheet, info in classification.items():
        role = info.get("role", "other")
        if role in ("comps", "sensitivity", "backup", "other"):
            continue
        conf = info.get("confidence", "low")
        by_role.setdefault(role, []).append((_CONF_RANK.get(conf, 2), sheet))

    nominated: dict[str, list[str]] = {}
    for role, entries in by_role.items():
        entries.sort(key=lambda x: x[0])  # high confidence first
        # Cap at 3 tabs per role. (2 was too tight — on BAC it dropped the One
        # Pager, where Levered IRR + Equity Multiple actually live, causing
        # returns 0/3.) 3 keeps the authoritative source plus two backups.
        nominated[role] = [name for _, name in entries][:3]

    # The catalog's source_primary lists (authored in the xlsx) use 'summary',
    # not the finer 'one_pager' role. Fold one_pager tabs into the 'summary'
    # nomination — FIRST, so the deal one-pager is the preferred summary source —
    # so metrics whose hierarchy says 'summary' actually read the one-pager.
    if nominated.get("one_pager"):
        merged = nominated["one_pager"] + [
            s for s in nominated.get("summary", []) if s not in nominated["one_pager"]
        ]
        nominated["summary"] = merged[:3]

    # Fallback authoritative pool — used when a metric's preferred roles map to
    # NO nominated tab (the classifier didn't confidently assign that role).
    # Without this, on complex models property metrics SKIP with "no target
    # sheets matched preferred list" and go silently missing. We surface the
    # highest-confidence one-pager/summary/inputs/sources_uses tabs as a catch-all.
    catch_all: list[str] = []
    for role in ("one_pager", "summary", "inputs", "sources_uses"):
        catch_all.extend(nominated.get(role, []))
    # de-dupe preserving order
    seen = set()
    nominated["_catch_all"] = [t for t in catch_all if not (t in seen or seen.add(t))][:4]
    return nominated


def effective_tier(
    sheet_name: str,
    name_based_tier: int,
    classification: dict[str, dict] | None,
) -> int:
    """
    Combine the deterministic name-based tier with GPT content classification.

    Rules (conservative — GPT corrects, doesn't override blindly):
      - If no classification available → use name-based tier unchanged.
      - If GPT role is HIGH confidence → trust the role-implied tier.
      - If GPT role is MEDIUM confidence → take the BETTER (lower) of the two
        tiers, but never promote a name-skipped sheet (99) into scanning unless
        GPT is high-confidence it's a real data sheet.
      - If GPT role is LOW confidence → keep name-based tier.

    This rescues mis-named real sheets (Tab3 → cash_flow) while not letting a
    shaky guess pull comps/sensitivity sheets back into the metric scan.
    """
    if not classification:
        return name_based_tier
    info = classification.get(sheet_name)
    if not info:
        return name_based_tier

    role_tier = info["implied_tier"]
    conf = info["confidence"]

    if conf == "high":
        return role_tier
    if conf == "medium":
        # Don't let a medium guess un-skip a sheet the name flagged as skip(99).
        if name_based_tier == 99 and role_tier != 99:
            return name_based_tier  # keep skipped — be conservative
        return min(name_based_tier, role_tier)
    # low confidence
    return name_based_tier
