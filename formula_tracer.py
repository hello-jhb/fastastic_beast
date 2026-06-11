"""
formula_tracer.py — Stage 2 of the redesign (2026-06-10).

After a human verifies the Audit Appendix Metrics (Stage 1), the verified cells
are trusted anchors. This module re-opens the workbook WITH FORMULAS
(`data_only=False`) and follows each anchor's formula references outward to
reach the rest of the model — the non-AAM catalog metrics (cash-flow detail,
returns waterfall, debt schedule, etc.).

This mirrors how an analyst actually reads a model: find the anchor numbers,
then trace the references to where the real detail lives. Empirically the
highest-value links are CROSS-SHEET references — e.g. `NOI = ='Annual CFs'!G46`
points straight at the cash-flow schedule, and `Levered IRR =
='Budget & Draw Schedule'!P104` at the returns tab.

Output is provenance-rich: every reached fact carries the anchor it came from,
the hop distance, the cell, the live value, and the formula chain. It does NOT
overwrite verified facts; it expands reach around them.
"""
from __future__ import annotations

import logging
import sys
from collections import deque
from pathlib import Path
from typing import Any

from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.cell import coordinate_to_tuple

from aam import AAM_METRIC_IDS
from flexible_extractor import normalize_text
from metric_catalog import load_metric_catalog

log = logging.getLogger("fb.formula_tracer")
if not log.handlers:
    _h = logging.StreamHandler(sys.stderr)
    _h.setFormatter(logging.Formatter("[fb.trace] %(message)s"))
    log.addHandler(_h)
    log.setLevel(logging.INFO)

FORMULA_TRACER_VERSION = "2026-06-10.1"

# Bounds so a pathological model can't blow up the trace.
_MAX_CELLS = 250            # total cells visited across the whole trace
_MAX_CELLS_PER_RANGE = 40   # cap when expanding a range like SUM(G4:G400)
_MAX_LABEL_SCAN_LEFT = 12   # how far left to scan for a row label


def trace_from_verified(
    file_path: str | Path,
    verified: dict[str, dict],
    max_hops: int = 2,
) -> dict[str, Any]:
    """
    Trace outward from verified AAM cells along formula references.

    Args:
      file_path: the workbook.
      verified:  { metric_name: record } with source_sheet / source_cell — the
                 trusted anchors (typically the confirmed Audit Appendix).
      max_hops:  how many reference hops to follow from each anchor.

    Returns:
      {
        "seeds":           <count of valid anchor cells>,
        "traced_cells":    [ {sheet, cell, hop, anchor, formula, value, label,
                              metric_id?, metric_name?}, ... ],
        "reached_metrics": { metric_id: {metric_name, value, source, label,
                              via_anchor, hop} },   # non-AAM only
      }
    """
    file_path = Path(file_path)
    import openpyxl
    wb_f = openpyxl.load_workbook(file_path, data_only=False)  # formulas
    wb_v = openpyxl.load_workbook(file_path, data_only=True)   # cached values
    sheets = set(wb_f.sheetnames)

    alias_index = _build_alias_index(
        load_metric_catalog(), exclude_ids=set(AAM_METRIC_IDS)
    )

    visited: dict[tuple[str, str], dict] = {}
    queue: deque[tuple[str, str, int, str]] = deque()

    seeds = 0
    for name, rec in verified.items():
        sheet, cell = rec.get("source_sheet"), rec.get("source_cell")
        if sheet in sheets and cell and rec.get("status") != "missing":
            queue.append((sheet, cell, 0, name))
            seeds += 1

    while queue and len(visited) < _MAX_CELLS:
        sheet, cell, hop, anchor = queue.popleft()
        key = (sheet, cell)
        if key in visited or sheet not in sheets:
            continue
        try:
            fval = wb_f[sheet][cell].value
            vval = wb_v[sheet][cell].value
            row, col = coordinate_to_tuple(cell)
        except Exception:
            continue

        is_formula = isinstance(fval, str) and fval.startswith("=")
        label = _nearest_label(wb_v[sheet], row, col)

        rec = {
            "sheet":   sheet,
            "cell":    cell,
            "hop":     hop,
            "anchor":  anchor,
            "formula": fval if is_formula else None,
            "value":   vval,
            "label":   label,
        }
        if label and hop > 0:  # don't re-label the anchors themselves
            m = _match_label_to_metric(label, alias_index)
            if m:
                rec["metric_id"] = m["metric_id"]
                rec["metric_name"] = m["metric_name"]
        visited[key] = rec

        if is_formula and hop < max_hops:
            for ref_sheet, ref_cell in _parse_refs(fval, sheet):
                if (ref_sheet, ref_cell) not in visited:
                    queue.append((ref_sheet, ref_cell, hop + 1, anchor))

    # Aggregate the non-AAM catalog metrics the trace reached (first hit wins —
    # the queue is breadth-first, so that's the closest cell to an anchor).
    reached: dict[str, dict] = {}
    for rec in visited.values():
        mid = rec.get("metric_id")
        if mid and mid not in reached:
            reached[mid] = {
                "metric_name": rec["metric_name"],
                "value":       rec["value"],
                "source":      f"{rec['sheet']}!{rec['cell']}",
                "label":       rec["label"],
                "via_anchor":  rec["anchor"],
                "hop":         rec["hop"],
            }

    log.info(
        "TRACE %s — %d seed(s), %d cell(s) visited, %d non-AAM metric(s) reached",
        file_path.name, seeds, len(visited), len(reached),
    )
    return {
        "seeds":           seeds,
        "traced_cells":    list(visited.values()),
        "reached_metrics": reached,
    }


# ---------------------------------------------------------------------------
# Reference parsing
# ---------------------------------------------------------------------------

def _parse_refs(formula: str, default_sheet: str) -> list[tuple[str, str]]:
    """Extract (sheet, cell) precedents from a formula string, expanding ranges."""
    out: list[tuple[str, str]] = []
    try:
        tokens = Tokenizer(formula).items
    except Exception:
        return out
    for t in tokens:
        if t.type == "OPERAND" and t.subtype == "RANGE":
            sheet, cellpart = _split_sheet(t.value, default_sheet)
            for c in _expand_range(cellpart):
                out.append((sheet, c))
    return out


def _split_sheet(ref: str, default_sheet: str) -> tuple[str, str]:
    """Split `'Annual CFs'!G46` / `Assumption!D23` / `D7` into (sheet, cellpart)."""
    if "!" in ref:
        sheet_part, cell_part = ref.rsplit("!", 1)
        sheet_part = sheet_part.strip()
        if sheet_part.startswith("'") and sheet_part.endswith("'"):
            sheet_part = sheet_part[1:-1]
        return sheet_part, cell_part
    return default_sheet, ref


def _expand_range(cell_part: str) -> list[str]:
    """Expand `G4:G8` → [G4..G8]; single cell → [cell]. Capped, $-tolerant."""
    cp = cell_part.replace("$", "")
    try:
        min_c, min_r, max_c, max_r = range_boundaries(cp)
    except Exception:
        return []
    if None in (min_c, min_r, max_c, max_r):
        return []  # whole-row/column ref (A:A) — not useful to enumerate
    cells: list[str] = []
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            cells.append(f"{get_column_letter(c)}{r}")
            if len(cells) >= _MAX_CELLS_PER_RANGE:
                return cells
    return cells


# ---------------------------------------------------------------------------
# Labels + catalog mapping
# ---------------------------------------------------------------------------

def _nearest_label(ws, row: int, col: int) -> str | None:
    """Nearest row label: scan left along the row for the first text cell."""
    for c in range(col - 1, max(0, col - _MAX_LABEL_SCAN_LEFT) - 1, -1):
        try:
            v = ws.cell(row=row, column=c).value
        except Exception:
            continue
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None


def _build_alias_index(catalog: list[dict], exclude_ids: set[str]) -> list[tuple[str, dict]]:
    """[(normalized_alias, metric)] for non-AAM metrics, longest alias first."""
    idx: list[tuple[str, dict]] = []
    for m in catalog:
        if m["metric_id"] in exclude_ids:
            continue
        for alias in m.get("aliases", []):
            a = normalize_text(alias)
            if a and len(a) >= 3:
                idx.append((a, m))
    idx.sort(key=lambda t: -len(t[0]))  # prefer more specific (longer) aliases
    return idx


def _match_label_to_metric(label: str, alias_index: list[tuple[str, dict]]) -> dict | None:
    """Map a traced cell's label to a non-AAM catalog metric (exact, then substring)."""
    lt = normalize_text(label)
    if not lt or len(lt) < 3:
        return None
    for a, m in alias_index:
        if a == lt:
            return m
    for a, m in alias_index:
        if a in lt:
            return m
    return None
