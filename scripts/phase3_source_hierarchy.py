"""
phase3_source_hierarchy.py — add per-metric source hierarchy + section grouping.

Phase 3 demotes the catalog from extractor to checklist+validator. For that, each
bounded metric needs:

  section          — which review section it belongs to (property / deal_basis /
                     leverage / returns / capex). The Section Reader extracts one
                     section at a time.
  source_primary   — ordered list of sheet ROLES that are authoritative for this
                     metric. The reader reads these first.
  source_forbidden — sheet ROLES that must NEVER supply this metric. The validator
                     REJECTS any value sourced from a forbidden role (this is what
                     would have killed "Total Project Cost = F&B sales from an
                     expense-analysis tab").

Roles come from re_knowledge.ALL_ROLES. Edit HIERARCHY below and re-run.
"""
from __future__ import annotations
from openpyxl import load_workbook

CATALOG_PATH = "Snapshot Metric.xlsx"

# metric_name -> {section, primary[roles], forbidden[roles]}
# secondary roles = anything not primary and not forbidden (reader falls back to them).
HIERARCHY: dict[str, dict] = {

    # ── Property section ──────────────────────────────────────────────
    "Asset Name": {
        "section": "property",
        "primary": ["summary", "inputs"],
        "forbidden": ["comps", "sensitivity", "market", "cash_flow", "debt"],
    },
    "Property Type": {
        "section": "property",
        "primary": ["summary", "inputs"],
        "forbidden": ["comps", "sensitivity", "market", "cash_flow", "debt"],
    },
    "Location": {
        "section": "property",
        "primary": ["summary", "inputs"],
        "forbidden": ["comps", "sensitivity", "market", "cash_flow", "debt"],
    },
    "Number of Properties": {
        "section": "property",
        "primary": ["summary", "inputs"],
        "forbidden": ["comps", "sensitivity", "market", "cash_flow"],
    },
    "Total SF": {
        "section": "property",
        "primary": ["summary", "inputs"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Total Units": {
        "section": "property",
        "primary": ["summary", "inputs"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Parking Spaces": {
        "section": "property",
        "primary": ["summary", "inputs"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Physical Occupancy": {
        "section": "property",
        "primary": ["summary", "inputs", "cash_flow"],
        "forbidden": ["comps", "sensitivity", "market"],
    },

    # ── Deal Basis section ────────────────────────────────────────────
    "Purchase Price": {
        "section": "deal_basis",
        "primary": ["summary", "sources_uses", "inputs"],
        "forbidden": ["cash_flow", "debt", "comps", "sensitivity", "market"],
    },
    "Total Project Cost": {
        "section": "deal_basis",
        "primary": ["sources_uses", "summary", "inputs"],
        "forbidden": ["cash_flow", "comps", "sensitivity", "market"],
    },
    "Net Operating Income (NOI)": {  # going-in (year 1)
        "section": "deal_basis",
        "primary": ["summary", "cash_flow", "inputs"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Going-in Cap Rate": {
        "section": "deal_basis",
        "primary": ["summary", "inputs"],
        "forbidden": ["comps", "sensitivity", "market", "cash_flow"],
    },
    "Hold Period": {
        "section": "deal_basis",
        "primary": ["summary", "inputs"],
        "forbidden": ["comps", "sensitivity", "market", "cash_flow"],
    },
    "Purchase Date": {
        "section": "deal_basis",
        "primary": ["summary", "inputs"],
        "forbidden": ["comps", "sensitivity", "market", "cash_flow"],
    },
    "Exit Date": {
        "section": "deal_basis",
        "primary": ["summary", "returns", "inputs"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Exit NOI": {
        "section": "deal_basis",
        "primary": ["summary", "returns", "cash_flow"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Exit Cap Rate": {
        "section": "deal_basis",
        "primary": ["summary", "returns", "inputs"],
        "forbidden": ["comps", "sensitivity", "market", "cash_flow"],
    },
    "Exit Value / Terminal Value": {
        "section": "deal_basis",
        "primary": ["summary", "returns"],
        "forbidden": ["comps", "sensitivity", "market"],
    },

    # ── Leverage section ──────────────────────────────────────────────
    "Original LTV": {
        "section": "leverage",
        "primary": ["debt", "sources_uses", "summary"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Debt Amount": {
        "section": "leverage",
        "primary": ["sources_uses", "debt", "summary"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Equity Invested": {
        "section": "leverage",
        "primary": ["sources_uses", "summary", "inputs"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Interest Rate": {
        "section": "leverage",
        "primary": ["debt", "inputs", "summary"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Interest Rate Spread": {
        "section": "leverage",
        "primary": ["debt", "inputs"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Interest Rate Cap": {
        "section": "leverage",
        "primary": ["debt", "inputs"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Loan Maturity": {
        "section": "leverage",
        "primary": ["debt", "inputs"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Interest-Only Period Remaining": {
        "section": "leverage",
        "primary": ["debt", "inputs"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "DSCR / Debt Coverage Ratio": {
        "section": "leverage",
        "primary": ["debt", "summary", "returns"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Debt Yield": {
        "section": "leverage",
        "primary": ["debt", "summary"],
        "forbidden": ["comps", "sensitivity", "market"],
    },

    # ── Returns section ───────────────────────────────────────────────
    "Levered IRR": {
        "section": "returns",
        "primary": ["returns", "summary"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Unlevered IRR": {
        "section": "returns",
        "primary": ["returns", "summary"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
    "Equity Multiple": {
        "section": "returns",
        "primary": ["returns", "summary"],
        "forbidden": ["comps", "sensitivity", "market"],
    },

    # ── CapEx section ─────────────────────────────────────────────────
    "CapEx Budget": {
        "section": "capex",
        "primary": ["capex", "sources_uses", "summary"],
        "forbidden": ["comps", "sensitivity", "market"],
    },
}


def main():
    wb = load_workbook(CATALOG_PATH)
    ws = wb.active
    header = [c.value for c in ws[1]]
    name_idx = header.index("metric_name") + 1

    for col in ("section", "source_primary", "source_forbidden"):
        if col not in header:
            ws.cell(row=1, column=ws.max_column + 1, value=col)
            header = [c.value for c in ws[1]]

    sec_idx  = header.index("section") + 1
    prim_idx = header.index("source_primary") + 1
    forb_idx = header.index("source_forbidden") + 1

    done = []
    for row in ws.iter_rows(min_row=2):
        name = row[name_idx - 1].value
        h = HIERARCHY.get(name)
        if not h:
            continue
        row[sec_idx - 1].value  = h["section"]
        row[prim_idx - 1].value = "; ".join(h["primary"])
        row[forb_idx - 1].value = "; ".join(h["forbidden"])
        done.append(name)

    wb.save(CATALOG_PATH)
    print(f"Source hierarchy applied to {len(done)} metrics.")
    by_section: dict[str, int] = {}
    for n in done:
        s = HIERARCHY[n]["section"]
        by_section[s] = by_section.get(s, 0) + 1
    for s, c in sorted(by_section.items()):
        print(f"  {s}: {c}")


if __name__ == "__main__":
    main()
