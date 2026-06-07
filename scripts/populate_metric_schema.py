"""
populate_metric_schema.py — add Phase 1 schema columns to Snapshot Metric.xlsx

Adds six new columns to every metric row:
  - unit              : USD | percent | ratio | months | years | count | sf | text
  - scale             : actual | thousands | millions | n/a
  - period            : at_close | year_1 | stabilized | exit | n/a
  - range_min         : minimum sane value (None if not bounded)
  - range_max         : maximum sane value (None if not bounded)
  - preferred_sheets  : ordered list (semicolon-separated) of sheet-name keywords
                       where this metric's canonical value lives

Also adds:
  - in_bounded_list   : True if this is one of the ~25 analyst-checklist metrics
                       that Phase 1 will route through the new candidate-based
                       extractor. Other metrics still work via the legacy path
                       until Phase 2.

Re-run this script after editing the SCHEMA_RULES dict below.
"""
from __future__ import annotations
from openpyxl import load_workbook

CATALOG_PATH = "Snapshot Metric.xlsx"

# Phase 1 — the bounded analyst-checklist metric list with schema.
# Keys must match the `metric_name` column in the Excel exactly.
# Anything not in this dict gets sensible defaults but is NOT bounded.
SCHEMA_RULES: dict[str, dict] = {

    # ── Property Detail ──────────────────────────────────────────────
    "Asset Name": {
        "unit": "text", "scale": "n/a", "period": "n/a",
        "range_min": None, "range_max": None,
        "preferred_sheets": "general info; investment summary; one pager; summary",
    },
    "Property Type": {
        "unit": "text", "scale": "n/a", "period": "n/a",
        "range_min": None, "range_max": None,
        "preferred_sheets": "general info; investment summary; one pager",
    },
    "Location": {
        "unit": "text", "scale": "n/a", "period": "n/a",
        "range_min": None, "range_max": None,
        "preferred_sheets": "general info; investment summary; one pager",
    },
    "Total SF": {
        "unit": "sf", "scale": "actual", "period": "stabilized",
        "range_min": 1_000, "range_max": 50_000_000,
        "preferred_sheets": "general info; investment summary; one pager",
    },
    "Total Units": {
        "unit": "count", "scale": "actual", "period": "stabilized",
        "range_min": 1, "range_max": 50_000,
        "preferred_sheets": "general info; investment summary; one pager",
    },
    "Parking Spaces": {
        "unit": "count", "scale": "actual", "period": "n/a",
        "range_min": 0, "range_max": 100_000,
        "preferred_sheets": "general info; investment summary",
    },
    "Physical Occupancy": {
        "unit": "ratio", "scale": "actual", "period": "at_close",
        "range_min": 0.0, "range_max": 1.0,
        "preferred_sheets": "general info; key uw; summary; one pager",
    },

    # ── Deal Basis ───────────────────────────────────────────────────
    "Purchase Price": {
        "unit": "USD", "scale": "actual", "period": "at_close",
        "range_min": 1_000_000, "range_max": 10_000_000_000,
        "preferred_sheets": "general info; investment summary; summary; sources & uses; key uw",
    },
    "Total Project Cost": {
        "unit": "USD", "scale": "actual", "period": "at_close",
        "range_min": 1_000_000, "range_max": 15_000_000_000,
        "preferred_sheets": "sources & uses; summary; general info; inputs; budget",
    },
    "Debt Amount": {
        "unit": "USD", "scale": "actual", "period": "at_close",
        "range_min": 1_000_000, "range_max": 10_000_000_000,
        "preferred_sheets": "sources & uses; inputs; debt; debt information; summary; general info",
    },
    "Equity Invested": {
        "unit": "USD", "scale": "actual", "period": "at_close",
        "range_min": 500_000, "range_max": 10_000_000_000,
        "preferred_sheets": "sources & uses; inputs; summary; general info; one pager",
    },
    "Net Operating Income (NOI)": {  # treated as Going-in NOI
        "unit": "USD", "scale": "actual", "period": "year_1",
        "range_min": 10_000, "range_max": 2_000_000_000,
        "preferred_sheets": "key uw; summary; cash flow; proforma; one pager",
    },
    "Going-in Cap Rate": {
        "unit": "ratio", "scale": "actual", "period": "at_close",
        "range_min": 0.01, "range_max": 0.20,
        "preferred_sheets": "key uw; summary; one pager",
    },
    "Hold Period": {
        "unit": "years", "scale": "actual", "period": "n/a",
        "range_min": 1, "range_max": 20,
        "preferred_sheets": "general info; investment summary; summary; key uw",
    },
    "Exit NOI": {
        "unit": "USD", "scale": "actual", "period": "exit",
        "range_min": 10_000, "range_max": 5_000_000_000,
        "preferred_sheets": "key uw; cash flow; summary",
    },
    "Exit Cap Rate": {
        "unit": "ratio", "scale": "actual", "period": "exit",
        "range_min": 0.01, "range_max": 0.20,
        "preferred_sheets": "key uw; summary; one pager",
    },
    "Exit Value / Terminal Value": {
        "unit": "USD", "scale": "actual", "period": "exit",
        "range_min": 1_000_000, "range_max": 50_000_000_000,
        "preferred_sheets": "key uw; summary; one pager; returns",
    },

    # ── Leverage ─────────────────────────────────────────────────────
    "Original LTV": {
        "unit": "ratio", "scale": "actual", "period": "at_close",
        "range_min": 0.10, "range_max": 0.95,
        "preferred_sheets": "debt; debt information; key uw; summary",
    },
    "Loan Maturity": {
        "unit": "months", "scale": "actual", "period": "n/a",
        "range_min": 1, "range_max": 600,
        "preferred_sheets": "debt; debt information",
    },
    "Interest-Only Period Remaining": {
        "unit": "months", "scale": "actual", "period": "n/a",
        "range_min": 0, "range_max": 240,
        "preferred_sheets": "debt; debt information",
    },
    "Interest Rate": {
        "unit": "ratio", "scale": "actual", "period": "at_close",
        "range_min": 0.005, "range_max": 0.25,
        "preferred_sheets": "debt; debt information; key uw; summary",
    },
    "DSCR / Debt Coverage Ratio": {
        "unit": "ratio", "scale": "actual", "period": "year_1",
        "range_min": 0.5, "range_max": 5.0,
        "preferred_sheets": "debt; debt information; key uw",
    },
    "Debt Yield": {
        "unit": "ratio", "scale": "actual", "period": "year_1",
        "range_min": 0.03, "range_max": 0.30,
        "preferred_sheets": "debt; debt information; key uw",
    },

    # ── Capital Spend ────────────────────────────────────────────────
    "CapEx Budget": {
        "unit": "USD", "scale": "actual", "period": "n/a",
        "range_min": 0, "range_max": 5_000_000_000,
        "preferred_sheets": "capex; sources & uses; budget; summary",
    },

    # ── Multi-Property Detection (Phase 1.5a) ────────────────────────
    "Number of Properties": {
        "unit": "count", "scale": "actual", "period": "n/a",
        "range_min": 1, "range_max": 100,
        "preferred_sheets": "general info; investment summary; one pager",
    },

    # ── Date Anchors for Hold Period (Phase 1.5a) ────────────────────
    "Purchase Date": {
        "unit": "date", "scale": "n/a", "period": "at_close",
        "range_min": None, "range_max": None,
        "preferred_sheets": "general info; investment summary; key uw",
    },
    "Exit Date": {
        "unit": "date", "scale": "n/a", "period": "exit",
        "range_min": None, "range_max": None,
        "preferred_sheets": "general info; investment summary; key uw; returns",
    },

    # ── Floating-Rate Debt Structure (Phase 1.5a) ────────────────────
    "Interest Rate Spread": {
        "unit": "ratio", "scale": "actual", "period": "at_close",
        "range_min": 0.001, "range_max": 0.10,
        "preferred_sheets": "debt; debt information",
    },
    "Interest Rate Cap": {
        "unit": "ratio", "scale": "actual", "period": "at_close",
        "range_min": 0.001, "range_max": 0.10,
        "preferred_sheets": "debt; debt information",
    },

    # ── Returns ──────────────────────────────────────────────────────
    "Levered IRR": {
        "unit": "ratio", "scale": "actual", "period": "n/a",
        "range_min": -0.20, "range_max": 0.80,
        "preferred_sheets": "key uw; returns; summary; one pager",
    },
    "Unlevered IRR": {
        "unit": "ratio", "scale": "actual", "period": "n/a",
        "range_min": -0.20, "range_max": 0.50,
        "preferred_sheets": "key uw; returns; summary; one pager",
    },
    "Equity Multiple": {
        "unit": "ratio", "scale": "actual", "period": "n/a",
        "range_min": 0.5, "range_max": 10.0,
        "preferred_sheets": "key uw; returns; summary; one pager",
    },
}


# Sensible defaults applied to every metric NOT in SCHEMA_RULES.
# These are placeholders — they let the system function without crashing,
# but the metric won't be routed through the bounded extractor.
DEFAULT_SCHEMA = {
    "unit": "USD",         # most non-bounded metrics are dollar values
    "scale": "actual",
    "period": "n/a",
    "range_min": None,
    "range_max": None,
    "preferred_sheets": "",
}


def main():
    wb = load_workbook(CATALOG_PATH)
    ws = wb.active
    header_row = [c.value for c in ws[1]]

    new_columns = [
        "unit",
        "scale",
        "period",
        "range_min",
        "range_max",
        "preferred_sheets",
        "in_bounded_list",
    ]
    name_col_idx = header_row.index("metric_name") + 1

    # Append new columns if they don't already exist
    for col_name in new_columns:
        if col_name not in header_row:
            ws.cell(row=1, column=ws.max_column + 1, value=col_name)
            header_row = [c.value for c in ws[1]]

    # Column indices (1-based for openpyxl)
    col_idx = {name: header_row.index(name) + 1 for name in new_columns}

    bounded_count = 0
    for row in ws.iter_rows(min_row=2):
        metric_name = row[name_col_idx - 1].value
        if not metric_name:
            continue

        if metric_name in SCHEMA_RULES:
            schema = SCHEMA_RULES[metric_name]
            in_bounded = True
            bounded_count += 1
        else:
            schema = DEFAULT_SCHEMA
            in_bounded = False

        for col_name in ("unit", "scale", "period", "range_min", "range_max", "preferred_sheets"):
            row[col_idx[col_name] - 1].value = schema.get(col_name)
        row[col_idx["in_bounded_list"] - 1].value = in_bounded

    wb.save(CATALOG_PATH)
    print(f"Updated {ws.max_row - 1} metric rows.")
    print(f"  Bounded list (Phase 1 routed): {bounded_count}")
    print(f"  Defaults: {ws.max_row - 1 - bounded_count}")


if __name__ == "__main__":
    main()
